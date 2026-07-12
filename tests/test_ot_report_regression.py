from __future__ import annotations

import importlib
import io
import json
import os
import sys
import types
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
import pytz
from openpyxl import load_workbook


FIXTURE_PATH = Path(__file__).resolve().parent / "fixtures" / "ot_regression" / "legacy_samples.json"
PROJECT_ROOT = Path(__file__).resolve().parents[1]

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


class DummyExpr:
    def __ge__(self, other):
        return self

    def __le__(self, other):
        return self


def _install_import_stubs(monkeypatch):
    os.environ.setdefault("DATABASE_URL", "postgres://user:pass@localhost/dbname")
    os.environ.setdefault("SECRET_KEY", "test")
    os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", "https://example.com/key.json")

    requests_stub = types.ModuleType("requests")
    requests_stub.get = lambda *_args, **_kwargs: SimpleNamespace(json=lambda: {})
    monkeypatch.setitem(sys.modules, "requests", requests_stub)

    main = types.ModuleType("app.main")
    main.app = SimpleNamespace(config={})
    main.db = SimpleNamespace(session=None)
    main.func = SimpleNamespace(timezone=lambda *_args, **_kwargs: DummyExpr())
    main.StaffPersonalInfo = object()
    main.StaffSpecialGroup = object()
    main.StaffShiftSchedule = object()
    main.StaffWorkLogin = object()
    main.StaffLeaveRequest = object()
    monkeypatch.setitem(sys.modules, "app.main", main)

    models = types.ModuleType("app.models")
    models.Org = object()
    monkeypatch.setitem(sys.modules, "app.models", models)

    forms = types.ModuleType("app.ot.forms")
    forms.pytz = pytz
    forms.OtPaymentAnnounceForm = object
    forms.OtCompensationRateForm = object
    forms.OtTimeSlotForm = object
    forms.OtDocumentApprovalForm = object
    forms.DateTimeRange = lambda lower=None, upper=None, bounds=None: SimpleNamespace(
        lower=lower,
        upper=upper,
        bounds=bounds,
    )
    forms.time_slots = []
    forms.create_ot_record_form = lambda *_args, **_kwargs: object
    forms.OtScheduleItemForm = object
    forms.OtScheduleForm = object
    monkeypatch.setitem(sys.modules, "app.ot.forms", forms)

    roles = types.ModuleType("app.roles")

    class Perm:
        def union(self, _other):
            return self

        def require(self):
            def decorator(fn):
                return fn

            return decorator

    roles.secretary_permission = Perm()
    roles.manager_permission = Perm()
    monkeypatch.setitem(sys.modules, "app.roles", roles)

    auth = types.ModuleType("pydrive.auth")

    class FakeCred:
        @classmethod
        def from_json_keyfile_dict(cls, _keyfile_dict, _scopes):
            return object()

    class FakeGA:
        def __init__(self, *_args, **_kwargs):
            self.credentials = None

    auth.ServiceAccountCredentials = FakeCred
    auth.GoogleAuth = FakeGA
    monkeypatch.setitem(sys.modules, "pydrive.auth", auth)

    drive = types.ModuleType("pydrive.drive")

    class FakeGD:
        def __init__(self, *_args, **_kwargs):
            pass

    drive.GoogleDrive = FakeGD
    monkeypatch.setitem(sys.modules, "pydrive.drive", drive)

    sys.modules.pop("app.ot.views", None)


@pytest.fixture
def ot_views(monkeypatch):
    _install_import_stubs(monkeypatch)
    views = importlib.import_module("app.ot.views")
    monkeypatch.setattr(views, "_is_external_account", lambda: False)
    return views


def _normalize_text(value):
    if value is None:
        return None
    return str(value).strip()


def _normalize_emp_id(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def _normalize_minutes(value):
    if value is None:
        return None
    return int(round(float(value)))


def _normalize_payment(value):
    if value is None:
        return None
    return round(float(value), 2)


def _minutes_to_display(minutes):
    if minutes is None:
        return None
    hours, mins = divmod(int(minutes), 60)
    if hours == 0 and mins == 0:
        return None
    return f"{hours}:{mins:02d}"


def _load_fixture():
    if not FIXTURE_PATH.exists():
        raise AssertionError(
            f"Missing regression fixture: {FIXTURE_PATH}. "
            "Run tests/fixtures/ot_regression/extract_legacy_ot_report.py first."
        )
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


def _build_dataframe(rows):
    frame = pd.DataFrame(rows)
    expected_columns = [
        "fullname",
        "sap",
        "position",
        "rate",
        "start",
        "end",
        "checkins",
        "checkouts",
        "late_checkin_display",
        "late_minutes",
        "early_checkout_display",
        "early_minutes",
        "work_minutes",
        "payment",
    ]
    return frame[expected_columns]


def _generate_workbook(ot_views, rows):
    buffer = io.BytesIO()
    frame = _build_dataframe(rows)
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        ot_views.write_ot_report_workbook(writer, frame, format="timesheet")
    buffer.seek(0)
    return load_workbook(buffer, data_only=True)


def _sheet_rows(worksheet):
    headers = [_normalize_text(cell.value) for cell in worksheet[1]]
    rows = []
    for raw in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in raw):
            continue
        rows.append(dict(zip(headers, raw)))
    return rows


def _pick_value(row, *candidates):
    for candidate in candidates:
        if candidate in row:
            return row[candidate]
    raise KeyError(f"Missing one of {candidates!r} in row: {row!r}")


def _parse_generated_timesheet_rows(worksheet):
    rows = []
    for item in _sheet_rows(worksheet):
        rows.append(
            {
                "fullname": _normalize_text(_pick_value(item, "ชื่อ", "fullname")),
                "sap": _normalize_emp_id(_pick_value(item, "รหัสบุคคล", "sap")),
                "position": _normalize_text(_pick_value(item, "ตำแหน่งงาน", "position")),
                "rate": _normalize_text(_pick_value(item, "อัตรา", "rate")),
                "start": _normalize_datetime(_pick_value(item, "เวลาเริ่มปฏิบัติงาน", "start")),
                "end": _normalize_datetime(_pick_value(item, "เวลาเลิกปฏิบัติงาน", "end")),
                "checkins": _normalize_datetime(_pick_value(item, "เวลาเข้างานจริง", "checkins")),
                "checkouts": _normalize_datetime(_pick_value(item, "เวลาออกงานจริง", "checkouts")),
                "late_checkin_display": _normalize_text(_pick_value(item, "late_checkin_display")),
                "late_minutes": _normalize_minutes(_pick_value(item, "late_minutes")),
                "early_checkout_display": _normalize_text(_pick_value(item, "early_checkout_display")),
                "early_minutes": _normalize_minutes(_pick_value(item, "early_minutes")),
                "work_minutes": _normalize_minutes(_pick_value(item, "จำนวนเวลาปฏิบัติงาน", "เวลาทำงาน", "work_minutes")),
                "payment": _normalize_payment(_pick_value(item, "จำนวนเงินที่ได้รับ", "payment")),
            }
        )
    return rows


def _parse_timesheet_fixture_rows(fixture_rows):
    rows = []
    for row in fixture_rows:
        rows.append(
            {
                "fullname": _normalize_text(row["fullname"]),
                "sap": _normalize_emp_id(row["sap"]),
                "position": _normalize_text(row["position"]),
                "rate": _normalize_text(row["rate"]),
                "start": _normalize_datetime(row["start"]),
                "end": _normalize_datetime(row["end"]),
                "checkins": _normalize_datetime(row["checkins"]),
                "checkouts": _normalize_datetime(row["checkouts"]),
                "late_checkin_display": _normalize_text(row["late_checkin_display"]),
                "late_minutes": _normalize_minutes(row["late_minutes"]),
                "early_checkout_display": _normalize_text(row["early_checkout_display"]),
                "early_minutes": _normalize_minutes(row["early_minutes"]),
                "work_minutes": _normalize_minutes(row["work_minutes"]),
                "payment": _normalize_payment(row["payment"]),
            }
        )
    return rows


def _row_key(row):
    return (
        row["sap"],
        row["fullname"],
        row["start"],
        row["end"],
    )


def _compare_row_sets(expected_rows, actual_rows, employee_label):
    expected_map = {_row_key(row): row for row in expected_rows}
    actual_map = {_row_key(row): row for row in actual_rows}

    assert set(actual_map) == set(expected_map), (
        f"{employee_label}: row keys differ\n"
        f"missing: {sorted(set(expected_map) - set(actual_map))}\n"
        f"extra: {sorted(set(actual_map) - set(expected_map))}"
    )

    for key in sorted(expected_map):
        expected = expected_map[key]
        actual = actual_map[key]
        for field in [
            "position",
            "rate",
            "checkins",
            "checkouts",
            "late_checkin_display",
            "late_minutes",
            "early_checkout_display",
            "early_minutes",
            "work_minutes",
            "payment",
        ]:
            assert actual[field] == expected[field], (
                f"{employee_label} {key[2]} -> {key[3]} field {field} differs\n"
                f"expected: {expected[field]!r}\n"
                f"actual:   {actual[field]!r}"
            )


def _compare_totals_sheet(worksheet, expected_rows):
    expected = {}
    for row in expected_rows:
        key = (row["fullname"], row["sap"])
        expected[key] = _minutes_to_display(row["work_minutes"])

    actual = {}
    for item in _sheet_rows(worksheet):
        key = (
            _normalize_text(_pick_value(item, "fullname", "ชื่อ", "ชื่อ - สกุล")),
            _normalize_emp_id(_pick_value(item, "sap", "รหัส sap", "รหัสบุคคล")),
        )
        actual[key] = _normalize_text(_pick_value(item, "work_minutes", "จำนวนเวลาปฏิบัติงาน"))

    assert set(actual) == set(expected), (
        f"total_minutes keys differ\n"
        f"missing: {sorted(set(expected) - set(actual))}\n"
        f"extra: {sorted(set(actual) - set(expected))}"
    )
    for key in sorted(expected):
        assert actual[key] == expected[key], (
            f"total_minutes differs for {key}\n"
            f"expected: {expected[key]!r}\n"
            f"actual:   {actual[key]!r}"
        )


def _compare_payment_sheet(worksheet, expected_rows):
    expected = {}
    for row in expected_rows:
        key = (row["fullname"], row["sap"])
        expected[key] = round(expected.get(key, 0.0) + _normalize_payment(row["payment"]), 2)

    actual = {}
    for item in _sheet_rows(worksheet):
        key = (
            _normalize_text(_pick_value(item, "fullname", "ชื่อ - สกุล", "ชื่อ")),
            _normalize_emp_id(_pick_value(item, "sap", "รหัส sap", "รหัสบุคคล")),
        )
        actual[key] = _normalize_payment(_pick_value(item, "payment", "จำนวนเงินที่ได้รับ"))

    assert set(actual) == set(expected), (
        f"total_payment keys differ\n"
        f"missing: {sorted(set(expected) - set(actual))}\n"
        f"extra: {sorted(set(actual) - set(expected))}"
    )
    for key in sorted(expected):
        assert actual[key] == pytest.approx(expected[key], abs=0.01), (
            f"total_payment differs for {key}\n"
            f"expected: {expected[key]!r}\n"
            f"actual:   {actual[key]!r}"
        )


def _group_by_employee(rows):
    grouped = {}
    for row in rows:
        key = (row["fullname"], row["sap"])
        grouped.setdefault(key, []).append(row)
    for employee_rows in grouped.values():
        employee_rows.sort(key=_row_key)
    return grouped


def test_ot_report_matches_legacy_sample_rows(ot_views):
    fixture = _load_fixture()
    selected_employees = fixture["selected_employees"]
    expected_rows_by_employee = {
        sap: _parse_timesheet_fixture_rows(rows)
        for sap, rows in fixture["rows_by_employee"].items()
    }

    all_expected_rows = []
    for rows in expected_rows_by_employee.values():
        all_expected_rows.extend(rows)

    workbook = _generate_workbook(ot_views, all_expected_rows)

    total_minutes_ws = workbook["total_minutes"]
    total_payment_ws = workbook["total_payment"]
    timesheet_ws = workbook["timesheet"]

    _compare_totals_sheet(total_minutes_ws, selected_employees)
    _compare_payment_sheet(total_payment_ws, all_expected_rows)

    generated_rows = _parse_generated_timesheet_rows(timesheet_ws)
    expected_grouped = _group_by_employee(all_expected_rows)
    actual_grouped = _group_by_employee(generated_rows)

    assert set(actual_grouped) == set(expected_grouped), (
        f"timesheet employees differ\n"
        f"missing: {sorted(set(expected_grouped) - set(actual_grouped))}\n"
        f"extra: {sorted(set(actual_grouped) - set(expected_grouped))}"
    )

    for employee_key in sorted(expected_grouped):
        expected_rows = expected_grouped[employee_key]
        actual_rows = actual_grouped[employee_key]
        assert len(actual_rows) == len(expected_rows), (
            f"{employee_key}: row count differs\n"
            f"expected: {len(expected_rows)}\n"
            f"actual:   {len(actual_rows)}"
        )
        _compare_row_sets(expected_rows, actual_rows, f"{employee_key[1]} {employee_key[0]}")


def test_ot_report_exports_handle_empty_dataframe(ot_views):
    empty = pd.DataFrame()
    cal_start = datetime(2026, 1, 1)
    cal_end = datetime(2026, 1, 2)

    report_buffer = ot_views.build_custom_ot_report_workbook(empty, cal_start, cal_end)
    finance_buffer = ot_views.build_finance_pdf(empty, cal_start, cal_end)

    assert report_buffer.getbuffer().nbytes > 0
    assert finance_buffer.getbuffer().nbytes > 0
