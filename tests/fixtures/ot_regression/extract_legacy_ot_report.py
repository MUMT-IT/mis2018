from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path("/Users/likitpreeyanon/Downloads/02-2024_ot_report_all_legacy.xlsx")
DEFAULT_OUTPUT = Path(__file__).with_name("legacy_samples.json")


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def _normalize_emp_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_datetime(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def _minutes_to_display(minutes: int | float | None) -> str | None:
    if minutes is None:
        return None
    minutes = int(round(float(minutes)))
    hours, mins = divmod(minutes, 60)
    if hours == 0 and mins == 0:
        return None
    return f"{hours}:{mins:02d}"


def load_timesheet_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["timesheet"]
    headers = [
        "fullname",
        "sap",
        "position",
        "rate",
        "start",
        "end",
        "checkins",
        "checkouts",
        "late_checkin_display",
        "late_minutes",
        "early_checkout_display",
        "early_minutes",
        "work_minutes",
        "payment",
    ]
    rows: list[dict[str, Any]] = []
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw_row))
        row["fullname"] = _normalize_text(row["fullname"])
        row["sap"] = _normalize_emp_id(row["sap"])
        row["position"] = _normalize_text(row["position"])
        row["rate"] = _normalize_text(row["rate"])
        row["start"] = _normalize_datetime(row["start"])
        row["end"] = _normalize_datetime(row["end"])
        row["checkins"] = _normalize_datetime(row["checkins"])
        row["checkouts"] = _normalize_datetime(row["checkouts"])
        row["late_checkin_display"] = _normalize_text(row["late_checkin_display"])
        row["early_checkout_display"] = _normalize_text(row["early_checkout_display"])
        row["late_minutes"] = int(row["late_minutes"] or 0)
        row["early_minutes"] = int(row["early_minutes"] or 0)
        row["work_minutes"] = int(row["work_minutes"] or 0)
        row["payment"] = round(float(row["payment"] or 0.0), 2)
        rows.append(row)
    return rows


def select_sample_employees(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    stats: dict[tuple[str | None, str | None], dict[str, Any]] = defaultdict(
        lambda: {"row_count": 0, "work_minutes": 0, "payment": 0.0}
    )
    for row in rows:
        key = (row["fullname"], row["sap"])
        stats[key]["row_count"] += 1
        stats[key]["work_minutes"] += row["work_minutes"]
        stats[key]["payment"] += row["payment"]

    ranked = {
        "many_rows": max(stats.items(), key=lambda item: (item[1]["row_count"], item[1]["work_minutes"], item[0][1] or "")),
        "few_rows": min(stats.items(), key=lambda item: (item[1]["row_count"], -item[1]["work_minutes"], item[0][1] or "")),
        "high_minutes": max(stats.items(), key=lambda item: (item[1]["work_minutes"], item[1]["row_count"], item[0][1] or "")),
        "high_payment": max(stats.items(), key=lambda item: (item[1]["payment"], item[1]["row_count"], item[0][1] or "")),
    }

    selected: list[dict[str, Any]] = []
    seen: set[tuple[str | None, str | None]] = set()
    for reason, (key, stat) in ranked.items():
        if key in seen:
            continue
        seen.add(key)
        fullname, sap = key
        selected.append(
            {
                "reason": reason,
                "fullname": fullname,
                "sap": sap,
                "row_count": stat["row_count"],
                "work_minutes": int(stat["work_minutes"]),
                "work_minutes_display": _minutes_to_display(stat["work_minutes"]),
                "payment": round(float(stat["payment"]), 2),
            }
        )
    return selected


def build_fixture_payload(workbook_path: Path) -> dict[str, Any]:
    rows = load_timesheet_rows(workbook_path)
    selected_employees = select_sample_employees(rows)
    selected_keys = {(item["fullname"], item["sap"]) for item in selected_employees}

    rows_by_employee: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (row["fullname"], row["sap"])
        if key not in selected_keys:
            continue
        rows_by_employee[row["sap"]].append(row)

    for sap_rows in rows_by_employee.values():
        sap_rows.sort(key=lambda row: (row["start"] or "", row["end"] or "", row["payment"], row["work_minutes"]))

    payload = {
        "source_file": str(workbook_path),
        "selected_employees": selected_employees,
        "rows_by_employee": rows_by_employee,
    }
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract regression samples from the legacy OT workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Path to the legacy workbook.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Where to write the JSON fixture.")
    args = parser.parse_args()

    payload = build_fixture_payload(args.input)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    print(f"Wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
