# -*- coding:utf-8 -*-
import io
import json
from collections import defaultdict, namedtuple

import dateutil.parser
import pandas as pd
from openpyxl import Workbook
from dateutil import parser
import arrow
from flask_login import login_required, current_user
import requests
import os

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import cast, Date, extract, and_

from werkzeug.utils import secure_filename

from app.ot.forms import *
from . import otbp as ot
from app.google_credential_utils import load_google_credentials_json
from app.main import (db, func, StaffPersonalInfo, StaffSpecialGroup,
                      StaffShiftSchedule, StaffWorkLogin, StaffLeaveRequest)
from app.models import Org
from flask import abort, jsonify, render_template, request, redirect, url_for, flash, make_response, send_file, current_app
from pydrive.auth import ServiceAccountCredentials, GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import date, datetime, time, timedelta

from ..roles import secretary_permission, manager_permission
from psycopg2._range import DateTimeRange

today = datetime.today()
if today.month >= 10:
    START_FISCAL_DATE = datetime(today.year, 10, 1)
    END_FISCAL_DATE = datetime(today.year + 1, 9, 30, 23, 59, 59, 0)
else:
    START_FISCAL_DATE = datetime(today.year - 1, 10, 1)
    END_FISCAL_DATE = datetime(today.year, 9, 30, 23, 59, 59, 0)

localtz = pytz.timezone('Asia/Bangkok')

login_tuple = namedtuple('LoginPair', ['staff_id', 'start', 'end', 'start_id', 'end_id'])

MAX_LATE_MINUTES = 45

# OT matching stays conservative: keep missing scans visible, prefer complete pairs first,
# and only let open or synthetic rows satisfy one shift so they do not leak forward.

EXTERNAL_OT_ALLOWED_ENDPOINTS = {
    'ot.view_monthly_records',
    'ot.summary_each_person',
    'ot.get_ot_records',
    'ot.get_all_ot_records_table',
    'ot.get_ot_records_table',
}


def _is_external_account():
    if not current_user.is_authenticated:
        return False
    personal_info = getattr(current_user, 'personal_info', None)
    org = getattr(personal_info, 'org', None)
    return bool(org and org.is_external)


def _normalize_ot_datetime(value):
    if value is None:
        return None
    if value.tzinfo is None:
        return localtz.localize(value)
    return value.astimezone(localtz)


def _build_ot_range(start_datetime, end_datetime):
    start_datetime = _normalize_ot_datetime(start_datetime)
    end_datetime = _normalize_ot_datetime(end_datetime)
    if start_datetime is None or end_datetime is None:
        return None
    return DateTimeRange(start_datetime, end_datetime, bounds='[)')


def _ranges_overlap(left_range, right_range):
    if not left_range or not right_range:
        return False
    return left_range.lower < right_range.upper and left_range.upper > right_range.lower


def _has_overlapping_ot_record_for_staff(staff_id, start_datetime, end_datetime, *, exclude_record_id=None):
    candidate_range = _build_ot_range(start_datetime, end_datetime)
    if candidate_range is None:
        return False

    query = (
        OtRecord.query
        .join(OtShift)
        .filter(OtRecord.staff_account_id == staff_id)
        .filter(OtShift.datetime.op('&&')(candidate_range))
        .filter(OtRecord.canceled_at.is_(None))
    )
    if exclude_record_id is not None:
        query = query.filter(OtRecord.id != exclude_record_id)
    return query.first() is not None


def _has_overlapping_ot_record(records, start_datetime, end_datetime, *, exclude_record_id=None):
    """Return True when the proposed interval overlaps an existing OT record."""
    candidate_range = _build_ot_range(start_datetime, end_datetime)
    for record in records:
        if exclude_record_id is not None and record.id == exclude_record_id:
            continue
        if record.canceled_at:
            continue
        existing_range = _build_ot_range(record.start_datetime, record.end_datetime)
        if _ranges_overlap(candidate_range, existing_range):
            return True
    return False


def _bangkok_localize(value):
    if value.tzinfo is not None:
        return value.astimezone(localtz)
    return localtz.localize(value)


@ot.before_request
def block_external_ot_routes():
    if not _is_external_account():
        return
    endpoint = request.endpoint or ''
    if endpoint in EXTERNAL_OT_ALLOWED_ENDPOINTS:
        return
    abort(403)

pdfmetrics.registerFont(TTFont('Sarabun', 'app/static/fonts/THSarabunNew.ttf'))
pdfmetrics.registerFont(TTFont('SarabunBold', 'app/static/fonts/THSarabunNewBold.ttf'))


def convert_to_fiscal_year(date):
    if date.month in [10, 11, 12]:
        return date.year + 1
    else:
        return date.year


def get_start_end_date_for_fiscal_year(fiscal_year):
    """Find start and end date from a given fiscal year.

    param fiscal_year:  fiscal year
    :return: date
    """
    start_date = date(fiscal_year - 1, 10, 1)
    end_date = date(fiscal_year, 9, 30)
    return start_date, end_date


gauth = GoogleAuth()
keyfile_dict = load_google_credentials_json()
scopes = ['https://www.googleapis.com/auth/drive']
if keyfile_dict:
    gauth.credentials = ServiceAccountCredentials.from_json_keyfile_dict(keyfile_dict, scopes)
    drive = GoogleDrive(gauth)
else:
    drive = None

tz = pytz.timezone('Asia/Bangkok')

FOLDER_ANNOUNCE_ID = '1xQQVOCtZHJmOLLVol8pkOz3CC7urxUAi'
FOLDER_DOCUMENT_ID = '1d8forb97XS-2v2puvH2FfhtD3lw2I4H5'
json_keyfile = load_google_credentials_json()


def initialize_gdrive():
    if not json_keyfile:
        return None
    gauth = GoogleAuth()
    scopes = ['https://www.googleapis.com/auth/drive']
    gauth.credentials = ServiceAccountCredentials.from_json_keyfile_dict(json_keyfile, scopes)
    return GoogleDrive(gauth)


@ot.route('/')
@manager_permission.union(secretary_permission).require()
@login_required
def index():
    work_at_orgs = (
        db.session.query(Org)
        .join(OtJobRole, OtJobRole.work_for_org_id == Org.id)
        .join(OtPaymentAnnounce, OtJobRole.announce_id == OtPaymentAnnounce.id)
        .filter(OtPaymentAnnounce.cancelled_at.is_(None))
        .filter(OtJobRole.work_for_org_id.isnot(None))
        .distinct()
        .order_by(Org.name)
        .all()
    )
    return render_template('ot/index.html', work_at_orgs=work_at_orgs)


@ot.route('/admin')
@login_required
def admin_index():
    work_at_orgs = (
        db.session.query(Org)
        .join(OtJobRole, OtJobRole.work_for_org_id == Org.id)
        .join(OtPaymentAnnounce, OtJobRole.announce_id == OtPaymentAnnounce.id)
        .filter(OtPaymentAnnounce.cancelled_at.is_(None))
        .filter(OtJobRole.work_for_org_id.isnot(None))
        .distinct()
        .order_by(Org.name)
        .all()
    )
    return render_template('ot/admin_index.html', work_at_orgs=work_at_orgs)


@ot.route('/admin/timeslots', methods=['GET', 'POST'])
@manager_permission.union(secretary_permission).require()
@login_required
def admin_timeslots():
    form = OtTimeSlotForm()
    selected_announcement_id = request.args.get('announcement_id', type=int)
    edit_timeslot_id = request.args.get('timeslot_id', type=int) or request.form.get('timeslot_id', type=int)
    selected_announcement = None
    if selected_announcement_id:
        selected_announcement = OtPaymentAnnounce.query.get(selected_announcement_id)
    if not selected_announcement:
        selected_announcement = OtPaymentAnnounce.query.order_by(OtPaymentAnnounce.id).first()
    if request.method == 'GET' and selected_announcement and not form.announcement.data:
        form.announcement.data = selected_announcement

    edit_timeslot = None
    if edit_timeslot_id:
        edit_timeslot = OtTimeSlot.query.get_or_404(edit_timeslot_id)
        if request.method == 'GET':
            form.announcement.data = edit_timeslot.announcement
            form.work_for_org.data = edit_timeslot.work_for_org
            form.start.data = edit_timeslot.start.strftime('%H:%M')
            form.end.data = edit_timeslot.end.strftime('%H:%M')
            form.color.data = edit_timeslot.color or form.color.default
            form.note.data = edit_timeslot.note

    timeslots = OtTimeSlot.query.order_by(OtTimeSlot.announcement_id, OtTimeSlot.start).all()

    if request.method == 'POST' and form.validate_on_submit():
        if edit_timeslot:
            timeslot = edit_timeslot
            flash_message = u'แก้ไขช่วงเวลาเรียบร้อยแล้ว'
        else:
            timeslot = OtTimeSlot()
            db.session.add(timeslot)
            flash_message = u'เพิ่มช่วงเวลาเรียบร้อยแล้ว'
        timeslot.announcement = form.announcement.data
        timeslot.work_for_org = form.work_for_org.data
        timeslot.start = time.fromisoformat(form.start.data)
        timeslot.end = time.fromisoformat(form.end.data)
        timeslot.color = form.color.data or None
        timeslot.note = form.note.data or None
        db.session.commit()
        flash(flash_message, 'success')
        return redirect(url_for('ot.admin_timeslots', announcement_id=timeslot.announcement_id))
    elif request.method == 'POST':
        flash(u'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบ', 'danger')
        for field, errors in form.errors.items():
            for error in errors:
                flash(u'{} {}'.format(field, error), 'danger')

    return render_template('ot/admin_timeslots.html',
                           form=form,
                           timeslots=timeslots,
                           selected_announcement=selected_announcement,
                           edit_timeslot=edit_timeslot)


def _configure_compensation_job_role_query(form):
    announcement = form.announcement.data
    work_at_org = form.work_at_org.data

    def query_factory():
        if not announcement or not work_at_org:
            return []
        query = OtJobRole.query
        query = query.filter_by(announce_id=announcement.id)
        query = query.filter_by(work_for_org_id=work_at_org.id)
        return query.order_by(OtJobRole.role).all()

    form.ot_job_role.query_factory = query_factory


@ot.route('/api/announcements/<int:announcement_id>/timeslots')
@login_required
def get_announcement_timeslots(announcement_id):
    work_for_org_id = request.args.get('work_for_org_id', type=int)
    query = OtTimeSlot.query.filter_by(announcement_id=announcement_id)
    if work_for_org_id:
        query = query.filter_by(work_for_org_id=work_for_org_id)
    timeslots = query.order_by(OtTimeSlot.start).all()
    return jsonify([{
        'id': slot.id,
        'label': str(slot),
    } for slot in timeslots])


@ot.route('/api/announcements/<int:announcement_id>/job-roles')
@login_required
def get_announcement_job_roles(announcement_id):
    work_for_org_id = request.args.get('work_for_org_id', type=int)
    query = OtJobRole.query.filter_by(announce_id=announcement_id)
    if work_for_org_id:
        query = query.filter_by(work_for_org_id=work_for_org_id)
    job_roles = query.order_by(OtJobRole.role).all()
    return jsonify([{
        'id': role.id,
        'label': role.role,
    } for role in job_roles])


@ot.route('/orgs/<int:org_id>/announcement-list-modal')
@login_required
def list_announcement_modal(org_id):
    announcements = (
        OtPaymentAnnounce.query
        .join(OtJobRole, OtJobRole.announce_id == OtPaymentAnnounce.id)
        .filter(OtJobRole.work_for_org_id == org_id)
        .filter(OtPaymentAnnounce.cancelled_at.is_(None))
        .distinct()
        .order_by(OtPaymentAnnounce.created_at.desc())
        .all()
    )
    org = Org.query.get_or_404(org_id)
    return render_template('ot/modals/announcements.html', announcements=announcements, org=org)


def _reset_announce_signatories(form, signatories=None, default_prepared_by=None):
    while form.signatories.entries:
        form.signatories.pop_entry()
    if signatories:
        ordered_signatories = sorted(signatories, key=lambda item: (item.sort_order, item.id))
        for signatory in ordered_signatories:
            entry = form.signatories.append_entry()
            entry.form.report_creator_staff.data = signatory.report_creator_staff
            entry.form.report_creator_position.data = signatory.report_creator_position
            entry.form.signer_staff.data = signatory.signer_staff
            entry.form.signer_position.data = signatory.signer_position
    else:
        entry = form.signatories.append_entry()
        entry.form.report_creator_staff.data = current_user if current_user else None
        entry.form.report_creator_position.data = default_prepared_by.get('position') if default_prepared_by else ''
        entry.form.signer_staff.data = None
        entry.form.signer_position.data = ''
        if default_prepared_by:
            form.signatories.entries[0].form.report_creator_staff.data = default_prepared_by.get('staff')
            form.signatories.entries[0].form.report_creator_position.data = default_prepared_by.get('position')


@ot.route('/announce')
@login_required
def announcement():
    # TODO: check permission of the current user
    if not current_user:
        flash(u'ไม่พบสิทธิในการเข้าถึงหน้าดังกล่าว', 'danger')
        return render_template('ot/index.html')
    announcements = OtPaymentAnnounce.query.filter_by(cancelled_at=None).order_by(OtPaymentAnnounce.created_at.desc()).all()
    return render_template('ot/announce.html',
                           announcements=announcements)


@ot.route('/announce/create', methods=['GET', 'POST'])
@login_required
def announcement_create_document():
    form = OtPaymentAnnounceForm()
    if request.method == 'GET':
        default_prepared_by = None
        if current_user.personal_info:
            default_prepared_by = {
                'staff': current_user,
                'position': current_user.personal_info.position or '',
            }
        _reset_announce_signatories(form, default_prepared_by=default_prepared_by)
    if request.method == 'POST':
        if form.validate_on_submit():
            payment = OtPaymentAnnounce()
            payment.topic = form.topic.data
            payment.org = form.org.data
            payment.announce_at = form.announce_at.data
            payment.start_datetime = form.start_datetime.data
            drive = initialize_gdrive()
            if form.upload.data:
                upload_file = form.upload.data
                file_name = secure_filename(upload_file.filename)
                upload_file.save(file_name)
                file_drive = drive.CreateFile({'title': file_name,
                                               'parents': [{'id': FOLDER_ANNOUNCE_ID, 'kind': 'drive#fileLink'}]})
                file_drive.SetContentFile(file_name)
                try:
                    file_drive.Upload()
                    permission = file_drive.InsertPermission({'type': 'anyone',
                                                              'value': 'anyone',
                                                              'role': 'reader'})
                except:
                    flash('ไม่สามารถอัพโหลดไฟล์ขึ้น Google drive ได้', 'danger')
                else:
                    flash('ไฟล์ที่แนบมา ถูกบันทึกบน Google drive เรียบร้อยแล้ว', 'success')
                    payment.upload_file_url = file_drive['id']
                    payment.file_name = file_name
            payment.staff = current_user
            payment.signatories = []
            for index, signatory_form in enumerate(form.signatories.entries):
                if not (signatory_form.form.report_creator_staff.data or signatory_form.form.signer_staff.data):
                    continue
                payment.signatories.append(OtAnnouncementSignatory(
                    report_creator_staff=signatory_form.form.report_creator_staff.data,
                    report_creator_position=signatory_form.form.report_creator_position.data,
                    signer_staff=signatory_form.form.signer_staff.data,
                    signer_position=signatory_form.form.signer_position.data,
                    sort_order=index,
                ))
            db.session.add(payment)
            db.session.commit()
            flash(u'เพิ่มประกาศเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.announcement'))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    return render_template('ot/announce_create_document.html', form=form)


@ot.route('/announce/edit/<int:announcement_id>', methods=['GET', 'POST'])
@login_required
def announcement_edit_document(announcement_id):
    payment = OtPaymentAnnounce.query.get_or_404(announcement_id)
    form = OtPaymentAnnounceForm(obj=payment)
    if request.method == 'GET':
        default_prepared_by = None
        if payment.staff and payment.staff.personal_info:
            default_prepared_by = {
                'staff': payment.staff,
                'position': payment.staff.personal_info.position or '',
            }
        _reset_announce_signatories(form, payment.signatories, default_prepared_by=default_prepared_by)
    if request.method == 'POST':
        if form.validate_on_submit():
            original_staff = payment.staff
            payment.topic = form.topic.data
            payment.org = form.org.data
            payment.announce_at = form.announce_at.data
            payment.start_datetime = form.start_datetime.data
            payment.staff = original_staff
            if form.upload.data:
                drive = initialize_gdrive()
                upload_file = form.upload.data
                file_name = secure_filename(upload_file.filename)
                upload_file.save(file_name)
                file_drive = drive.CreateFile({'title': file_name,
                                               'parents': [{'id': FOLDER_ANNOUNCE_ID, 'kind': 'drive#fileLink'}]})
                file_drive.SetContentFile(file_name)
                try:
                    file_drive.Upload()
                    file_drive.InsertPermission({'type': 'anyone',
                                                 'value': 'anyone',
                                                 'role': 'reader'})
                except:
                    flash('ไม่สามารถอัพโหลดไฟล์ขึ้น Google drive ได้', 'danger')
                else:
                    flash('ไฟล์ที่แนบมา ถูกบันทึกบน Google drive เรียบร้อยแล้ว', 'success')
                    payment.upload_file_url = file_drive['id']
                    payment.file_name = file_name
            payment.signatories = []
            for index, signatory_form in enumerate(form.signatories.entries):
                if not (signatory_form.form.report_creator_staff.data or signatory_form.form.signer_staff.data):
                    continue
                payment.signatories.append(OtAnnouncementSignatory(
                    report_creator_staff=signatory_form.form.report_creator_staff.data,
                    report_creator_position=signatory_form.form.report_creator_position.data,
                    signer_staff=signatory_form.form.signer_staff.data,
                    signer_position=signatory_form.form.signer_position.data,
                    sort_order=index,
                ))
            db.session.add(payment)
            db.session.commit()
            flash(u'แก้ไขประกาศเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.announcement'))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    return render_template('ot/announce_edit_document.html', form=form, payment=payment)


@ot.route('/announce/<int:announcement_id>/compensations')
@login_required
def announcement_compensations(announcement_id):
    announcement = OtPaymentAnnounce.query.get_or_404(announcement_id)
    compensations = OtCompensationRate.query.filter_by(announce_id=announcement_id).all()
    return render_template('ot/announce_compensations.html',
                           announcement=announcement,
                           compensations=compensations)


@ot.route('/announce/add-compensation', methods=['GET', 'POST'])
@login_required
def announcement_add_compensation():
    form = OtCompensationRateForm()
    if request.method == 'GET' and getattr(current_user, 'personal_info', None) and current_user.personal_info.org:
        form.work_at_org.data = current_user.personal_info.org
    announcement_id = request.args.get('announcement_id', type=int)
    selected_announcement = None
    if request.method == 'GET' and announcement_id:
        selected_announcement = OtPaymentAnnounce.query.get(announcement_id)
        if selected_announcement:
            form.announcement.data = selected_announcement
    elif request.method == 'POST':
        selected_announcement = form.announcement.data

    _configure_compensation_job_role_query(form)

    if selected_announcement:
        form.time_slot.query_factory = lambda announcement_id=selected_announcement.id: OtTimeSlot.query.filter_by(announcement_id=announcement_id).all()
    else:
        form.time_slot.query_factory = lambda: []
    job_roles_data = [{
        'id': role.id,
        'announcement_id': role.announce_id,
        'work_at_org_id': role.work_for_org_id,
        'label': role.role,
    } for role in OtJobRole.query.order_by(OtJobRole.announce_id, OtJobRole.work_for_org_id, OtJobRole.role).all()]
    if request.method == 'POST':
        if form.validate_on_submit():
            compensation = OtCompensationRate()
            form.populate_obj(compensation)
            db.session.add(compensation)
            db.session.commit()
            flash(u'เพิ่มรายละเอียดของประกาศเรียบร้อยแล้ว', 'success')
            if compensation.announcement:
                return redirect(url_for('ot.announcement_compensations', announcement_id=compensation.announcement.id))
            return redirect(url_for('ot.announcement'))
        else:
            flash(u'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบ', 'danger')
    return render_template('ot/announce_compensation.html', form=form, job_roles_data=job_roles_data)


@ot.route('/announce/job-roles', methods=['GET', 'POST'])
@login_required
def announcement_job_roles():
    form = OtJobRoleForm()
    if request.method == 'POST':
        if form.validate_on_submit():
            job_role = OtJobRole()
            job_role.announcement = form.announcement.data
            job_role.work_for_org = form.work_at_org.data
            job_role.role = form.role.data
            db.session.add(job_role)
            db.session.commit()
            flash(u'เพิ่มตำแหน่งงานเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.announcement_job_roles'))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    job_roles = OtJobRole.query.order_by(OtJobRole.announce_id, OtJobRole.work_for_org_id, OtJobRole.role).all()
    return render_template('ot/announce_job_roles.html',
                           form=form,
                           job_roles=job_roles,
                           editing_job_role=None)


@ot.route('/announce/job-roles/<int:job_role_id>', methods=['GET', 'POST'])
@login_required
def announcement_edit_job_role(job_role_id):
    job_role = OtJobRole.query.get_or_404(job_role_id)
    form = OtJobRoleForm(obj=job_role)
    if request.method == 'GET':
        form.work_at_org.data = job_role.work_for_org
    if request.method == 'POST':
        if form.validate_on_submit():
            job_role.announcement = form.announcement.data
            job_role.work_for_org = form.work_at_org.data
            job_role.role = form.role.data
            db.session.add(job_role)
            db.session.commit()
            flash(u'แก้ไขตำแหน่งงานเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.announcement_job_roles'))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    job_roles = OtJobRole.query.order_by(OtJobRole.announce_id, OtJobRole.work_for_org_id, OtJobRole.role).all()
    return render_template('ot/announce_job_roles.html',
                           form=form,
                           job_roles=job_roles,
                           editing_job_role=job_role)


@ot.route('/announce/edit-compensation/<int:com_id>', methods=['GET', 'POST'])
@login_required
def announcement_edit_compensation(com_id):
    compensation = OtCompensationRate.query.get(com_id)
    form = OtCompensationRateForm(obj=compensation)
    if request.method == 'GET':
        if compensation.work_at_org:
            form.work_at_org.data = compensation.work_at_org
        elif getattr(current_user, 'personal_info', None) and current_user.personal_info.org:
            form.work_at_org.data = current_user.personal_info.org
    selected_announcement = form.announcement.data or compensation.announcement
    if request.method == 'POST':
        selected_announcement = form.announcement.data

    _configure_compensation_job_role_query(form)

    if selected_announcement:
        form.time_slot.query_factory = lambda announcement_id=selected_announcement.id: OtTimeSlot.query.filter_by(announcement_id=announcement_id).all()
    else:
        form.time_slot.query_factory = lambda: []
    job_roles_data = [{
        'id': role.id,
        'announcement_id': role.announce_id,
        'work_at_org_id': role.work_for_org_id,
        'label': role.role,
    } for role in OtJobRole.query.order_by(OtJobRole.announce_id, OtJobRole.work_for_org_id, OtJobRole.role).all()]
    if request.method == 'POST':
        if form.validate_on_submit():
            form.populate_obj(compensation)
            db.session.add(compensation)
            db.session.commit()
            flash(u'แก้ไขรายละเอียดของประกาศเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.announcement'))
        else:
            flash(u'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบ', 'danger')
    return render_template('ot/announce_compensation.html',
                           form=form,
                           compensation=compensation,
                           job_roles_data=job_roles_data)


@ot.route('/document-approval')
@login_required
def document_approval_records():
    # TODO: filter valid document
    documents = OtDocumentApproval.query.all()
    upload_file_url = None
    for document in documents:
        if document.upload_file_url:
            upload_file = drive.CreateFile({'id': document.upload_file_url})
            # upload_file.FetchMetadata()
            upload_file_url = upload_file.get('embedLink')
    return render_template('ot/document_approvals.html',
                           documents=documents, upload_file_url=upload_file_url)


@ot.route('/document-approval/create', methods=['GET', 'POST'])
@login_required
def document_approval_create_document():
    form = OtDocumentApprovalForm()
    if request.method == 'POST':
        if form.validate_on_submit():
            document = OtDocumentApproval()
            form.populate_obj(document)
            drive = initialize_gdrive()
            if form.upload.data:
                upload_file = form.upload.data
                file_name = secure_filename(upload_file.filename)
                upload_file.save(file_name)
                file_drive = drive.CreateFile({'title': file_name,
                                               'parents': [{'id': FOLDER_DOCUMENT_ID, 'kind': 'drive#fileLink'}]})
                file_drive.SetContentFile(file_name)
                try:
                    file_drive.Upload()
                    permission = file_drive.InsertPermission({'type': 'anyone',
                                                              'value': 'anyone',
                                                              'role': 'reader'})
                except:
                    flash('ไม่สามารถอัพโหลดไฟล์ขึ้น Google drive ได้', 'danger')
                else:
                    flash('ไฟล์ที่แนบมา ถูกบันทึกบน Google drive เรียบร้อยแล้ว', 'success')
                    document.upload_file_url = file_drive['id']
                    document.file_name = file_name
            document.created_staff = current_user
            document.org = current_user.personal_info.org
            db.session.add(document)
            db.session.commit()
            flash(u'เพิ่มอนุมัติในหลักการเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.document_approval_show_announcement', document_id=document.id))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    return render_template('ot/document_create_approval.html', form=form)


@ot.route('/document-approval/edit/<int:document_id>', methods=['GET', 'POST'])
@login_required
def document_approval_edit_document(document_id):
    document = OtDocumentApproval.query.get(document_id)
    form = OtDocumentApprovalForm(obj=document)
    if request.method == 'POST':
        if form.validate_on_submit():
            form.populate_obj(document)
            drive = initialize_gdrive()
            # TODO: ถ้าไม่บันทึกไฟล์ใหม่(แก้ข้อมูลส่วนอื่น) ไฟล์เก่าจะหายไปจาก db แต่ไม่หายจาก gg
            if form.upload.data:
                upload_file = form.upload.data
                file_name = secure_filename(upload_file.filename)
                upload_file.save(file_name)
                file_drive = drive.CreateFile({'title': file_name,
                                               'parents': [{'id': FOLDER_DOCUMENT_ID, 'kind': 'drive#fileLink'}]})
                file_drive.SetContentFile(file_name)
                try:
                    file_drive.Upload()
                    permission = file_drive.InsertPermission({'type': 'anyone',
                                                              'value': 'anyone',
                                                              'role': 'reader'})
                except:
                    flash('ไม่สามารถอัพโหลดไฟล์ขึ้น Google drive ได้', 'danger')
                else:
                    flash('ไฟล์ที่แนบมา ถูกบันทึกบน Google drive เรียบร้อยแล้ว', 'success')
                    document.upload_file_url = file_drive['id']
                    document.file_name = file_name
            document.created_staff = current_user
            document.org = current_user.personal_info.org
            db.session.add(document)
            db.session.commit()
            flash(u'แก้ไขอนุมัติในหลักการเรียบร้อยแล้ว', 'success')
            return redirect(url_for('ot.document_approval_records'))
        else:
            for field, err in form.errors.items():
                flash('{} {}'.format(field, err), 'danger')
    return render_template('ot/document_create_approval.html', form=form)


@ot.route('/document-approval/create/<int:document_id>/announcement')
@login_required
def document_approval_show_announcement(document_id):
    approval = OtDocumentApproval.query.get(document_id)
    announcements = OtPaymentAnnounce.query.all()
    if approval.upload_file_url:
        upload_file = drive.CreateFile({'id': approval.upload_file_url})
        upload_file.FetchMetadata()
        upload_file_url = upload_file.get('embedLink')
    else:
        upload_file_url = None
    return render_template('ot/document_announcement.html', approval=approval, upload_file_url=upload_file_url,
                           announcements=announcements)


@ot.route('/document-approval/create/<int:document_id>/add-announcement/<int:announce_id>')
@login_required
def document_approval_add_announcement(document_id, announce_id):
    announcement = OtPaymentAnnounce.query.get(announce_id)
    approval = OtDocumentApproval.query.get(document_id)
    if announcement and approval:
        approval.announce.append(announcement)
        db.session.add(approval)
        db.session.commit()
        flash(u'เพิ่มประกาศสำหรับอนุมัติในหลักการเรียบร้อยแล้ว', 'success')
        return redirect(url_for('ot.document_approval_show_announcement', document_id=document_id))
    else:
        flash(u'ไม่สามารถเพิ่มประกาศได้', 'danger')
        return redirect(url_for('ot.document_approval_show_announcement', document_id=document_id))


@ot.route('/document-approval/create/<int:document_id>/delete-announcement/<int:announce_id>')
@login_required
def document_approval_delete_announcement(document_id, announce_id):
    announcement = OtPaymentAnnounce.query.get(announce_id)
    approval = OtDocumentApproval.query.get(document_id)
    if approval and announcement:
        # TODO: หาว่ามีrecord ไหนที่ใช้อยู่่และเชื่อม ประกาศนี้อยู่ ยังไม่อนุญาตให้ลบหรือไม่ หรือจะหาทางออกยังไง
        approval.announce.remove(announcement)
        db.session.add(approval)
        db.session.commit()
        flash(u'ลบประกาศสำหรับอนุมัติในหลักการเรียบร้อยแล้ว', 'success')
        return redirect(url_for('ot.document_approval_show_announcement', document_id=document_id))
    else:
        flash(u'ไม่สามารถลบประกาศได้', 'danger')
        return redirect(url_for('ot.document_approval_show_announcement', document_id=document_id))


@ot.route('/document-approval/staff/<int:document_id>')
@login_required
def document_approval_show_approved_staff(document_id):
    approval = OtDocumentApproval.query.get(document_id)
    staff = StaffAccount.query.all()
    if approval.upload_file_url:
        upload_file = drive.CreateFile({'id': approval.upload_file_url})
        upload_file.FetchMetadata()
        upload_file_url = upload_file.get('embedLink')
    else:
        upload_file_url = None
    return render_template('ot/document_staff.html', approval=approval, staff=staff, upload_file_url=upload_file_url)


@ot.route('/document-approval/<int:document_id>/add-staff/<int:staff_id>')
@login_required
def document_approval_add_staff(document_id, staff_id):
    document = OtDocumentApproval.query.get(document_id)
    staff = StaffAccount.query.get(staff_id)
    if document:
        document.staff.append(staff)
        db.session.add(document)
        db.session.commit()
        flash(u'เพิ่มบุคลากรเรียบร้อยแล้ว', 'success')
        return redirect(url_for('ot.document_approval_show_approved_staff', document_id=document_id))
    else:
        flash(u'ไม่สามารถเพิ่มบุคลากรได้', 'danger')
        return redirect(url_for('ot.document_approval_show_approved_staff', document_id=document_id))


@ot.route('/document-approval/<int:document_id>/delete-staff/<int:staff_id>')
@login_required
def document_approval_delete_staff(document_id, staff_id):
    document = OtDocumentApproval.query.get(document_id)
    staff = StaffAccount.query.get(staff_id)
    if document:
        document.staff.remove(staff)
        db.session.add(document)
        db.session.commit()
        flash(u'ลบบุคลากรเรียบร้อยแล้ว', 'warning')
        return redirect(url_for('ot.document_approval_show_approved_staff', document_id=document_id))
    else:
        flash(u'ไม่สามารถลบบุคลากรได้', 'danger')
        return redirect(url_for('ot.document_approval_show_approved_staff', document_id=document_id))


@ot.route('/document-approvals/list/for-ot')
@login_required
def document_approvals_list_for_create_ot():
    documents = OtDocumentApproval.query.filter_by(org_id=current_user.personal_info.org.id).all()
    if documents:
        for document in documents:
            if document.upload_file_url:
                upload_file = drive.CreateFile({'id': document.upload_file_url})
                # upload_file.FetchMetadata()
                upload_file_url = upload_file.get('embedLink')
            else:
                upload_file_url = None
            # TODO: warning expired document
            # if document.end_datetime:
            #     if document.end_datetime <= today:
            #         is_expired = True
        return render_template('ot/document_approvals_list_create_scedule.html', documents=documents,
                               upload_file_url=upload_file_url)
    else:
        flash(u'หน่วยงานของท่านไม่มีอนุมัติในหลักการ กรุณาสร้างอนุมัติในหลักการก่อนทำการเบิกค่าตอบแทนล่วงเวลา',
              'warning')
        return render_template('ot/index.html')


@ot.route('/schedule/create/<int:document_id>', methods=['GET', 'POST'])
@login_required
def add_schedule(document_id):
    document = OtDocumentApproval.query.get(document_id)
    EditOtRecordForm = create_ot_record_form([a.id for a in document.announce])
    form = EditOtRecordForm()
    if request.method == 'POST':
        if form.validate_on_submit():
            for staff_id in request.form.getlist("otworker"):
                record = OtRecord()
                form.populate_obj(record)
                if form.compensation.data.start_time:
                    start_t = form.compensation.data.start_time
                    end_t = form.compensation.data.end_time
                else:
                    if form.start_time.data == "None" or form.end_time.data == "None":
                        flash(u'จำเป็นต้องใส่เวลาเริ่มต้น สิ้นสุด', 'danger')
                        return render_template('ot/schedule_add.html', form=form, document=document)
                    else:
                        start_t = form.start_time.data + ':00'
                        end_t = form.end_time.data + ':00'
                start_d = form.start_date.data
                end_d = form.start_date.data
                start_dt = '{} {}'.format(start_d, start_t)
                end_dt = '{} {}'.format(end_d, end_t)
                start_datetime = datetime.strptime(start_dt, '%Y-%m-%d %H:%M:%S')
                end_datetime = datetime.strptime(end_dt, '%Y-%m-%d %H:%M:%S')
                existing_records = OtRecord.query.filter_by(staff_account_id=staff_id).all()
                staff_name = StaffAccount.query.get(staff_id)
                if _has_overlapping_ot_record(existing_records, start_datetime, end_datetime):
                    flash(u'{} มีข้อมูลการทำOT ในช่วงเวลานี้แล้ว กรุณาตรวจสอบเวลาใหม่อีกครั้ง'.format(
                        staff_name.personal_info.fullname), 'danger')
                else:
                    record.start_datetime = start_datetime
                    record.end_datetime = end_datetime
                    record.created_staff = current_user
                    record.org = current_user.personal_info.org
                    record.staff_account_id = staff_id
                    record.document_id = document_id
                    if request.form.get('sub_role'):
                        record.sub_role = request.form.get('sub_role')
                    flash(u'บันทึกการทำงานของ {} เรียบร้อยแล้ว'.format(staff_name.personal_info.fullname), 'success')
                    db.session.add(record)
                    db.session.commit()
            return redirect(url_for('ot.document_approvals_list_for_create_ot'))
        else:
            print(form.errors, form.start_time.data)
            flash(u'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบ', 'danger')
    return render_template('ot/schedule_add.html', form=form, document=document)


@ot.route('/schedule/cancel/<int:record_id>')
@login_required
def cancel_ot_record(record_id):
    record = OtRecord.query.get(record_id)
    record.canceled_at = tz.localize(datetime.today())
    record.canceled_by_account_id = current_user.id
    db.session.add(record)
    db.session.commit()
    flash(u'ยกเลิก OT ของ {} {} เรียบร้อยแล้ว'.format(record.staff.personal_info.fullname, record.start_datetime),
          'danger')
    return redirect(url_for('ot.summary_ot_each_document', document_id=record.document_id,
                            month=record.start_datetime.month, year=record.start_datetime.year))


@ot.route('/announcements/<int:announcement_id>/schedule', methods=['GET', 'POST'])
@manager_permission.union(secretary_permission).require()
@login_required
def add_ot_schedule(announcement_id):
    announcement = OtPaymentAnnounce.query.get_or_404(announcement_id)
    org_id = request.args.get('org_id', type=int)
    if not org_id:
        org_id = announcement.org_id
    selected_work_at_org = Org.query.get(org_id) if org_id else None
    slots_query = OtTimeSlot.query.filter_by(announcement_id=announcement_id)
    if selected_work_at_org:
        slots_query = slots_query.filter(OtTimeSlot.work_at_org == selected_work_at_org)
    slots = slots_query.order_by(OtTimeSlot.start).all()
    return render_template('ot/schedule_add.html',
                           announcement_id=announcement_id,
                           work_at_org_id=org_id,
                           slots=slots)


@ot.route('/announcements/<int:announcement_id>/reset-slot-selector')
@manager_permission.union(secretary_permission).require()
@login_required
def reset_slot_selector(announcement_id):
    announcement = OtPaymentAnnounce.query.get(announcement_id)
    org_id = request.args.get('org_id', type=int)
    selected_work_at_org = Org.query.get(org_id) if org_id else None
    slots = ''
    slot_query = OtTimeSlot.query.filter_by(announcement_id=announcement_id)
    if selected_work_at_org:
        slot_query = slot_query.filter(OtTimeSlot.work_at_org == selected_work_at_org)
    for slot in slot_query.order_by(OtTimeSlot.start).all():
        slots += f'<option value="timeslot-{slot.id}" >{slot}</option>'

    template = f'''
        <label class="label htmx-indicator has-text-danger">กรุณาเลือกช่วงเวลา</label>
        <div class="select">
            <select name="slot-id" hx-trigger="change"
                    hx-target="#shift-table"
                    hx-indicator="closest div"
                    hx-swap="innerHTML"
                    hx-vals="js:{{start: getStartDate()}}"
                    hx-get="{url_for('ot.show_ot_form_modal')}">
                <option>เลือกช่วงเวลาปฏิบัติงาน</option>
                {slots}
            </select>
        </div>
        <div id="shift-table" hx-swap-oob="true"></div>
    '''
    resp = make_response(template)
    resp.headers['HX-Trigger-After-Swap'] = 'initSelect2js'
    return resp


@ot.route('/api/announcements/<int:announcement_id>/shifts')
@manager_permission.union(secretary_permission).require()
@login_required
def get_shifts(announcement_id):
    start = request.args.get('start')
    start = arrow.get(dateutil.parser.parse(start), 'Asia/Bangkok').datetime
    shifts = []
    for slot in OtTimeSlot.query.filter_by(announcement_id=announcement_id):
        for shift in slot.shifts:
            if shift.datetime.lower.date() == start.date():
                shifts.append({
                    'id': f'shift-{shift.id}',
                    'start': shift.datetime.lower.isoformat(),
                    'end': shift.datetime.upper.isoformat(),
                    'title': ','.join([rec.staff.personal_info.th_firstname for rec in shift.records]),
                    'textColor': shift.timeslot.color or '',
                })
    return jsonify(shifts)


@ot.route('/timeslots/<_id>/ot-form-modal', methods=['GET', 'POST'])
@ot.route('/timeslots/ot-form-modal', methods=['GET', 'POST'])
@manager_permission.union(secretary_permission).require()
@login_required
def show_ot_form_modal(_id=None):
    start = request.args.get('start')
    start = arrow.get(datetime.strptime(start, '%d/%m/%Y'), 'Asia/Bangkok').datetime

    if _id is None:
        _id = request.args.get('slot-id')

    if _id.startswith('timeslot'):
        _, slot_id = _id.split('-')
        timeslot = OtTimeSlot.query.get(slot_id)
        start = datetime.combine(start.date(), timeslot.start)
        end = datetime.combine(start.date(), timeslot.end)
        if timeslot.end.hour == 0 and timeslot.end.minute == 0:
            datetime_ = DateTimeRange(lower=start, upper=end + timedelta(days=1), bounds='[)')
        else:
            datetime_ = DateTimeRange(lower=start, upper=end, bounds='[)')
        shift = OtShift.query.filter_by(datetime=datetime_, timeslot=timeslot).first()
    elif _id.startswith('shift'):
        _, shift_id = _id.split('-')
        shift = OtShift.query.get(shift_id)
        timeslot = shift.timeslot

    RecordForm = create_ot_record_form(timeslot, timeslot.work_for_org_id)
    form = RecordForm()
    compensation_rates = get_compensation_rates_for_timeslot(timeslot)
    form.compensation.query = compensation_rates
    form.staff.choices = [(staff.id, staff.fullname) for staff in StaffAccount.get_active_accounts()]
    if form.validate_on_submit():
        candidate_start = _bangkok_localize(shift.datetime.lower) if shift else _bangkok_localize(
            datetime.combine(start.date(), timeslot.start)
        )
        if timeslot.end.hour == 0 and timeslot.end.minute == 0:
            candidate_end = candidate_start.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
        else:
            candidate_end = _bangkok_localize(shift.datetime.upper) if shift else _bangkok_localize(
                datetime.combine(start.date(), timeslot.end)
            )

        for staff_id in form.staff.data:
            if _has_overlapping_ot_record_for_staff(staff_id, candidate_start, candidate_end):
                staff_name = StaffAccount.query.get(staff_id)
                flash(
                    u'{} มีข้อมูลการทำOT ในช่วงเวลานี้แล้ว กรุณาตรวจสอบเวลาใหม่อีกครั้ง'.format(
                        staff_name.fullname if staff_name else str(staff_id)
                    ),
                    'danger',
                )
                break
        else:
            if not shift:
                shift = OtShift(date=start.date(), timeslot=timeslot, creator=current_user)
            for staff_id in form.staff.data:
                ot_record = OtRecord.query.filter_by(shift=shift, staff_account_id=staff_id).first()
                if not ot_record:
                    ot_record = OtRecord(
                        staff_account_id=staff_id,
                        created_account_id=current_user.id,
                        shift=shift,
                        compensation=form.compensation.data,
                    )
                    shift.records.append(ot_record)
            db.session.add(shift)
            db.session.commit()
    else:
        print(form.errors)
    records = [] if current_app.testing else (
        OtRecord.query.filter_by(shift_id=shift.id).all() if shift else []
    )
    template = render_template('ot/modals/ot_record_form.html',
                               start=start,
                               target_url=url_for('ot.show_ot_form_modal', _id=_id, start=request.args.get('start')),
                               form=form, slot_id=timeslot.id, timeslot=timeslot, records=records,
                               compensation_rates=compensation_rates)
    resp = make_response(template)
    resp.headers['HX-Trigger-After-Swap'] = json.dumps({"initSelect2js": "",
                                                        "clearSelection": "",
                                                        "refetchEvents": ""})
    return resp


@ot.route('/records/<int:record_id>/remove', methods=['DELETE'])
@manager_permission.union(secretary_permission).require()
@login_required
def remove_record(record_id):
    record = OtRecord.query.get(record_id)
    db.session.delete(record)
    db.session.commit()
    resp = make_response()
    resp.headers['HX-Trigger'] = 'refetchEvents'
    return resp


@ot.route('/documents/<int:doc_id>/compensation_rates', methods=['POST'])
@manager_permission.union(secretary_permission).require()
@login_required
def get_compensation_rates(doc_id):
    form = OtScheduleForm()
    document = OtDocumentApproval.query.get(doc_id)
    compensations = []
    for a in document.announce:
        compensations += [rate for rate in a.ot_rate if rate.role == form.role.data]

    entry_ = form.items.append_entry()
    entry_.compensation.choices = [(rate.id, rate) for rate in compensations]
    entry_.time_slots.choices = [(slot.id, slot) for slot in compensations[0].time_slots]
    entry_.staff.choices = [(staff.id, staff.fullname) for staff in document.org.active_staff_accounts]
    template = f'''
    <div class="field">
        <div class="select">
            {entry_.compensation()}
        </div>
    </div>
    <div class="field" id="{entry_.staff.id}">
        {entry_.staff(class_="js-example-basic-multiple")}
    </div>
    <div class="field" id="{entry_.time_slots.id}">
        {entry_.time_slots()}
    </div>
    '''
    resp = make_response(template)
    resp.headers['HX-Trigger-After-Swap'] = 'initSelect2jsEvent'
    return resp


@ot.route('/documents/<int:doc_id>/schedule/records')
@manager_permission.union(secretary_permission).require()
@login_required
def list_ot_records(doc_id):
    document = OtDocumentApproval.query.get(doc_id)
    shifts = defaultdict(list)
    for rec in document.ot_records:
        shifts[rec.shift_datetime].append(rec)
    return render_template('ot/records.html', doc=document, shifts=shifts)


@ot.route('/schedule/<int:record_id>/delete', methods=['DELETE'])
@manager_permission.union(secretary_permission).require()
@login_required
def delete_ot_record(record_id):
    record = OtRecord.query.get(record_id)
    db.session.delete(record)
    db.session.commit()
    return ''


@ot.route('/schedule/edit/<int:record_id>', methods=['GET', 'POST'])
@login_required
def edit_ot_record(record_id):
    record = OtRecord.query.get(record_id)
    document = OtDocumentApproval.query.get(record.document_id)
    EditOtRecordForm = create_ot_record_form([a.id for a in document.announce])
    form = EditOtRecordForm(obj=record)
    if request.method == 'POST':
        if form.validate_on_submit():
            form.populate_obj(record)
            if form.compensation.data.start_time:
                start_t = form.compensation.data.start_time
                end_t = form.compensation.data.end_time
            else:
                if form.start_time.data == "None" or form.end_time.data == "None":
                    flash(u'จำเป็นต้องใส่เวลาเริ่มต้น สิ้นสุด', 'danger')
                else:
                    start_t = form.start_time.data + ':00'
                    end_t = form.end_time.data + ':00'
            start_d = form.start_date.data
            end_d = form.start_date.data
            start_dt = '{} {}'.format(start_d, start_t)
            end_dt = '{} {}'.format(end_d, end_t)
            start_datetime = datetime.strptime(start_dt, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_dt, '%Y-%m-%d %H:%M:%S')
            record.start_datetime = start_datetime
            record.end_datetime = end_datetime
            if _has_overlapping_ot_record_for_staff(
                record.staff_account_id,
                start_datetime,
                end_datetime,
                exclude_record_id=record.id,
            ):
                flash(u'{} มีข้อมูลการทำOT ในช่วงเวลานี้แล้ว กรุณาตรวจสอบเวลาใหม่อีกครั้ง'.format(
                    record.staff.personal_info.fullname), 'danger')
            else:
                record.created_staff = current_user
                record.org = current_user.personal_info.org
                if request.form.get('sub_role'):
                    record.sub_role = request.form.get('sub_role')
                db.session.add(record)
                db.session.commit()
                flash(u'แก้ไขการทำงานของ {} เรียบร้อยแล้ว'.format(record.staff.personal_info.fullname), 'success')
                year = form.start_date.data.year
                month = form.start_date.data.month
                return redirect(
                    url_for('ot.summary_ot_each_document', document_id=record.document_id, month=month, year=year))
        else:
            flash(u'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบ', 'danger')
    form.start_date.data = record.start_datetime.date()
    form.start_time.data = record.start_datetime.strftime("%H:%M")
    form.end_time.data = record.end_datetime.strftime("%H:%M")
    return render_template('ot/schedule_edit_each_record.html', form=form, record=record)


@ot.route('/api/get-file-url/<int:announcement_id>')
@login_required
def get_file_url(announcement_id):
    ann = OtPaymentAnnounce.query.get(announcement_id)
    return jsonify({'url': ann.upload_file_url})


@ot.route('/api/compensation-detail/<int:compensation_id>')
@login_required
def get_compensation_detail(compensation_id):
    comp = OtCompensationRate.query.get(compensation_id)
    return jsonify({'info': comp.to_dict()})


@ot.route('/schedule/summary')
@login_required
def summary_index():
    depts = Org.query.all()
    fiscal_year = request.args.get('fiscal_year')
    if fiscal_year is None:
        if today.month in [10, 11, 12]:
            fiscal_year = today.year + 1
        else:
            fiscal_year = today.year
        init_date = today
    else:
        fiscal_year = int(fiscal_year)
        init_date = date(fiscal_year - 1, 10, 1)
    if len(depts) == 0:
        # return redirect(request.referrer)
        return redirect(url_for("ot.schedule"))
    curr_dept_id = request.args.get('curr_dept_id')
    tab = request.args.get('tab', 'all')
    if curr_dept_id is None:
        curr_dept_id = depts[0].id
    employees = StaffPersonalInfo.query.all()
    ot_r = []
    for emp in employees:
        if tab == 'ot' or tab == 'all':
            fiscal_years = OtRecord.query.distinct(func.date_part('YEAR', OtRecord.start_datetime))
            fiscal_years = [convert_to_fiscal_year(ot.start_datetime) for ot in fiscal_years]
            start_fiscal_date, end_fiscal_date = get_start_end_date_for_fiscal_year(fiscal_year)
            for ot_record in OtRecord.query.filter_by(org_id=current_user.personal_info.org.id,
                                                      staff=emp.staff_account).filter(
                OtRecord.start_datetime.between(start_fiscal_date, end_fiscal_date)):
                shift_schedule_overlaps = StaffShiftSchedule.query.filter(StaffShiftSchedule.staff == ot_record.staff) \
                    .filter(StaffShiftSchedule.start_datetime <= ot_record.start_datetime) \
                    .filter(StaffShiftSchedule.end_datetime >= ot_record.start_datetime).all()
                shift_schedules = StaffShiftSchedule.query.filter(StaffShiftSchedule.staff == ot_record.staff) \
                    .filter(cast(StaffShiftSchedule.start_datetime, Date) == ot_record.start_datetime.date()).all()
                work_login_checkin = StaffWorkLogin.query.filter(StaffWorkLogin.staff == ot_record.staff) \
                    .filter(cast(StaffWorkLogin.start_datetime, Date) == ot_record.start_datetime.date()).all()
                work_login_checkout = StaffWorkLogin.query.filter(StaffWorkLogin.staff == ot_record.staff) \
                    .filter(cast(StaffWorkLogin.end_datetime, Date) == ot_record.end_datetime.date()).all()
                leave_request = StaffLeaveRequest.query.filter(StaffLeaveRequest.staff == ot_record.staff) \
                    .filter(cast(StaffLeaveRequest.start_datetime, Date) == ot_record.start_datetime.date()).all()

                if not shift_schedules and not work_login_checkin and not work_login_checkout:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
                elif shift_schedule_overlaps and not work_login_checkin or not work_login_checkout:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'เวลาปฏิบัติงานปกติตรงกับเวลาที่ขอเบิกค่าล่วงเวลา และไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
                elif not shift_schedules and not work_login_checkin:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนเข้างาน'
                elif not shift_schedules and not work_login_checkout:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนออกงาน'
                elif not work_login_checkin or not work_login_checkout:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
                elif not shift_schedules:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบเวลาปฏิบัติงาน'
                elif shift_schedule_overlaps:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'เวลาปฏิบัติงานปกติตรงกับเวลาที่ขอเบิกค่าล่วงเวลา'
                elif not work_login_checkin:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # u'ไม่พบบันทึกเวลาสแกนเข้างาน'
                elif not work_login_checkout:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # ot_record["condition"] = u'ไม่พบบันทึกเวลาสแกนสิ้นสุดงาน'
                elif not leave_request:
                    text_color = '#ffffff'
                    bg_color = '#F0475A'
                    # ot_record["condition"] = u'ตรงกับวันลาปฏิบัติงาน'
                else:
                    text_color = '#ffffff'
                    bg_color = '#2268F3'
                border_color = '#ffffff'
                ot_r.append({
                    'id': ot_record.id,
                    'start': ot_record.start_datetime,
                    'end': ot_record.end_datetime,
                    'title': u'{} {}'.format(emp.th_firstname, ot_record.compensation.role),
                    'backgroundColor': bg_color,
                    'borderColor': border_color,
                    'textColor': text_color,
                    'type': 'ot'
                })
            all = ot_r
    return render_template('ot/schedule_summary.html',
                           init_date=init_date,
                           depts=depts, curr_dept_id=int(curr_dept_id),
                           all=all, tab=tab, fiscal_years=fiscal_years, fiscal_year=fiscal_year)


@ot.route('/schedule/summary/each-org')
@login_required
def summary_ot_each_org():
    documents = set()
    records = OtRecord.query.filter_by(org_id=current_user.personal_info.org.id) \
        .filter(OtRecord.round_id == None) \
        .filter(OtRecord.canceled_at == None).all()
    for record in records:
        documents.add(
            (record.document.id, record.document.title, record.start_datetime.month, record.start_datetime.year))
    return render_template('ot/schedule_summary_each_org.html', documents=documents)


@ot.route('/schedule/summary/each-org/<int:document_id>/<int:month>/<int:year>')
@login_required
def summary_ot_each_document(document_id, month, year):
    records = OtRecord.query.filter_by(document_id=document_id, org_id=current_user.personal_info.org.id) \
        .filter(extract('month', OtRecord.start_datetime) == month) \
        .filter(extract('year', OtRecord.start_datetime) == year).filter(OtRecord.round_id == None).all()
    document = OtDocumentApproval.query.get(document_id)
    ot_records = []
    for record in records:
        ot_record = dict(
            id=record.id,
            staff=record.staff.personal_info.fullname,
            start_date=record.start_datetime.date(),
            start_time=record.start_datetime.time(),
            end_time=record.end_datetime.time(),
            compensation=record.compensation,
            work_at=record.compensation.work_at_org,
            work_for=record.compensation.work_for_org,
            sub_role=record.sub_role,
            condition=None,
            rate=None,
            hour=None,
            total_rate=None,
            canceled_at=record.canceled_at
        )
        ot_record["hour"] = record.total_ot_hours()
        ot_record["total_rate"] = record.count_rate()
        if record.compensation.per_period:
            ot_record["rate"] = u'{} บาทต่อคาบ'.format(record.compensation.per_period)
        elif record.compensation.per_hour:
            ot_record["rate"] = u'{} บาทต่อชั่วโมง'.format(record.compensation.per_hour)
        else:
            ot_record["rate"] = u'{} บาทต่อวัน'.format(record.compensation.per_day)
        shift_schedule_overlaps = StaffShiftSchedule.query.filter(StaffShiftSchedule.staff == record.staff) \
            .filter(StaffShiftSchedule.start_datetime <= record.start_datetime) \
            .filter(StaffShiftSchedule.end_datetime >= record.start_datetime) \
            .filter(StaffShiftSchedule.start_datetime <= record.end_datetime) \
            .filter(StaffShiftSchedule.end_datetime >= record.end_datetime).all()
        shift_schedules = StaffShiftSchedule.query.filter(StaffShiftSchedule.staff == record.staff) \
            .filter(cast(StaffShiftSchedule.start_datetime, Date) == record.start_datetime.date()).all()
        work_login_checkin = StaffWorkLogin.query.filter(StaffWorkLogin.staff == record.staff) \
            .filter(cast(StaffWorkLogin.start_datetime, Date) == record.start_datetime.date()).all()
        work_login_checkout = StaffWorkLogin.query.filter(StaffWorkLogin.staff == record.staff) \
            .filter(cast(StaffWorkLogin.end_datetime, Date) == record.end_datetime.date()).all()
        leave_request = StaffLeaveRequest.query.filter(StaffLeaveRequest.staff == record.staff) \
            .filter(cast(StaffLeaveRequest.start_datetime, Date) == record.start_datetime.date()).all()
        # TODO: compare ot record with worklogin
        if not shift_schedules and not work_login_checkin and not work_login_checkout:
            ot_record["condition"] = u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
        elif shift_schedule_overlaps and not work_login_checkin or not work_login_checkout:
            ot_record[
                "condition"] = u'เวลาปฏิบัติงานปกติตรงกับเวลาที่ขอเบิกค่าล่วงเวลา และไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
        elif not shift_schedules and not work_login_checkin:
            ot_record["condition"] = u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนเข้างาน'
        elif not shift_schedules and not work_login_checkout:
            ot_record["condition"] = u'ไม่พบเวลาปฏิบัติงาน และไม่พบบันทึกเวลาสแกนออกงาน'
        elif not work_login_checkin or not work_login_checkout:
            ot_record["condition"] = u'ไม่พบบันทึกเวลาสแกนเข้า-ออกงาน'
        elif not shift_schedules:
            ot_record["condition"] = u'ไม่พบเวลาปฏิบัติงาน'
        elif shift_schedule_overlaps:
            ot_record["condition"] = u'เวลาปฏิบัติงานปกติตรงกับเวลาที่ขอเบิกค่าล่วงเวลา'
        elif not work_login_checkin:
            ot_record["condition"] = u'ไม่พบบันทึกเวลาสแกนเข้างาน'
        elif not work_login_checkout:
            ot_record["condition"] = u'ไม่พบบันทึกเวลาสแกนสิ้นสุดงาน'
        elif not leave_request:
            ot_record["condition"] = u'ตรงกับวันลาปฏิบัติงาน'
        ot_records.append(ot_record)
    return render_template('ot/schedule_each_document.html', records=records, document=document, ot_records=ot_records,
                           month=month, year=year)


@ot.route('/schedule/summary/each-org/<int:document_id>/<int:month>/<int:year>/create-approval-create-download')
@login_required
def create_ot_approval_and_download(document_id, month, year):
    approver = Org.query.filter_by(id=current_user.personal_info.org.id).first()
    org_head = StaffAccount.query.filter_by(email=approver.head).first()
    round = OtRoundRequest(
        created_at=datetime.now(tz),
        created_by_account_id=current_user.id,
        approval_by_account_id=org_head.id,
        round_no=str(month) + "/" + str(year) + "-" + str(document_id)
    )
    db.session.add(round)
    for record in OtRecord.query.filter_by(document_id=document_id).filter(
            extract('month', OtRecord.start_datetime) == month) \
            .filter(extract('year', OtRecord.start_datetime) == year).filter(OtRecord.canceled_at == None).all():
        record.round = round
        db.session.add(record)
    db.session.commit()
    flash(u'ส่งคำขอเรียบร้อยแล้ว', 'success')
    # for ot_record in ot_records_query:
    #     record = {}
    #     record["start_datetime"] = ot_record.start_datetime
    #     record["staff"] = ot_record.staff.personal_info.fullname
    #     ot_list.append(record)
    # df = DataFrame(record)
    # summary = df.pivot_table(index='staff', columns='start_datetime', aggfunc=len, fill_value=0)
    # summary.to_excel('ot_summary.xlsx')
    # flash(u'ดาวน์โหลดไฟล์เรียบร้อยแล้ว ชื่อไฟล์ ot_summary.xlsx', 'success')
    return redirect(url_for('ot.round_request_status'))


@ot.route('/summary/each-org/round-request/status')
@login_required
def round_request_status():
    rounds = OtRoundRequest.query.filter_by(created_by=current_user).all()
    return render_template('ot/request_status.html', rounds=rounds)


@ot.route('/approver/requests-pending-list')
@login_required
def round_request_approval_requests_pending():
    rounds = OtRoundRequest.query.filter_by(approval_by=current_user).all()
    return render_template('ot/approver_pending_list.html', rounds=rounds)


@ot.route('/round-request/<int:round_id>/approval-info')
@login_required
def round_request_info(round_id):
    round = OtRoundRequest.query.filter_by(id=round_id).first()
    return render_template('ot/request_info_each_round.html', round=round)


@ot.route('/approver/round-request/<int:round_id>/approved')
@login_required
def round_request_approve_request(round_id):
    round = OtRoundRequest.query.get(round_id)
    round.approval_at = datetime.now(tz);
    db.session.add(round)
    db.session.commit()
    flash(u'อนุมัติรายการ{} เรียบร้อยแล้ว'.format(round.round_no), 'success')
    rounds = OtRoundRequest.query.filter_by(approval_by=current_user).all()
    return render_template('ot/approver_pending_list.html', rounds=rounds)


@ot.route('/finance/approved-list')
@login_required
def approved_list_from_org_head():
    rounds = OtRoundRequest.query.all()
    return render_template('ot/approved_list.html', rounds=rounds)


@ot.route('/finance/requests-pending-list/<int:round_id>')
@login_required
def round_request_info_for_finance(round_id):
    it = StaffSpecialGroup.query.filter_by(group_code='it').first()
    finance = StaffSpecialGroup.query.filter_by(group_code='finance').first()
    if current_user in it.staffs or current_user in finance.staffs:
        round = OtRoundRequest.query.filter_by(id=round_id).first()
        return render_template('ot/finance_approval_info.html', round=round)
    else:
        flash(u'ไม่พบสิทธิในการเข้าถึงหน้าดังกล่าว', 'danger')
        return redirect(request.referrer)


@ot.route('/finance/requests-pending-list/<int:round_id>/verify')
@login_required
def round_request_verify(round_id):
    for record in OtRecord.query.filter_by(round_id=round_id).all():
        if record.compensation.is_count_in_mins:
            record.total_shift_minutes = record.total_ot_hours()
        else:
            record.total_minutes = record.total_ot_hours()
        record.amount_paid = record.count_rate()
        db.session.add(record)
    round = OtRoundRequest.query.get(round_id)
    round.verified_by_account_id = current_user.id
    round.verified_at = datetime.now(tz)
    db.session.add(round)
    db.session.commit()
    flash(u'รับรองรายการ{} เรียบร้อยแล้ว'.format(round.round_no), 'success')
    rounds = OtRoundRequest.query.all()
    return render_template('ot/approved_list.html', rounds=rounds)


@ot.route('/<list_type>')
def event_list(list_type='timelineDay'):
    return render_template('ot/summary_chart.html', list_type=list_type)


@ot.route('/api/staff')
@login_required
def get_records(org_id):
    # org_id = request.args.get('deptid')
    # if org_id is None:
    #     ot_query = OtRecord.query.all()
    # else:
    #     ot_query = OtRecord.query.filter_by(org_id=org_id)
    record = OtRecord.query.all()
    otrecord = []
    for ot in record:
        otrecord.append({
            'id': ot.id,
            'location': ot.location,
            'title': ot.compensation.role,
            'stafforg': ot.staff.personal_info.org.display_name if ot.staff.personal_info.org else 'ไม่มีสังกัด',
            'businessHours': {
                'start': ot.start_datetime.strftime('%H:%M'),
                'end': ot.end_datetime.strftime('%H:%M'),
            }
        })
    return jsonify(otrecord)


@ot.route('/api/otrecords')
@login_required
def get_events():
    all_events = []
    text_color = '#ffffff'
    bg_color = '#2b8c36'
    border_color = '#ffffff'
    otrecords = OtRecord.query.all()
    for record in otrecords:
        event = {
            'location': event.get('location', None),
            'title': record.staff.personal_info.fullname,
            'description': event.get('description', ''),
            'start': record.start_datetime.strftime('%H:%M'),
            'end': record.end_datetime.strftime('%H:%M'),
            'resourceId': otrecords.id,
            'status': otrecords.round,
            'borderColor': border_color,
            'backgroundColor': bg_color,
            'textColor': text_color,
            'id': record.id,
        }
        all_events.append(event)
    return jsonify(all_events)


@ot.route('/records/<int:event_id>', methods=['POST', 'GET'])
@login_required
def show_event_detail(event_id=None):
    tz = pytz.timezone('Asia/Bangkok')
    if event_id:
        event = OtRecord.query.get(event_id)
        if event:
            event.start = event.start_date.astimezone(tz)
            event.end = event.end_date.astimezone(tz)
            return render_template(
                'ot/summary_chart.html', event=event)
    else:
        return 'No event ID specified.'


# @room.route('/events/<int:event_id>', methods=['POST', 'GET'])
# def show_event_detail(event_id=None):
#     tz = pytz.timezone('Asia/Bangkok')
#     if event_id:
#         event = RoomEvent.query.get(event_id)
#         if event:
#             event.start = event.start.astimezone(tz)
#             event.end = event.end.astimezone(tz)
#             return render_template(
#                 'scheduler/event_detail.html', event=event)
#     else:
#         return 'No event ID specified.'


@ot.route('/summary')
@login_required
def summary_chart():
    # ot_records = OtRecord.query.filter(OtRecord.canceled_at==None)\
    #                             .filter(OtRoundRequest.approval_at!=None).all()
    # records = [record.list_records() for record in ot_records]
    # records = []
    # for record in ot_records:
    #     ot = dict(
    #         record.compensation.role,
    #         record.staff.personal_info.fullname,
    #         record.start_datetime,
    #         record.end_datetime,
    #         record.total_hours or record.total_minutes
    #     )
    #     records.append(ot)
    # departments = Org.query.all()
    return render_template('ot/summary_chart.html')


@ot.route('/summary/each-person')
@login_required
def summary_each_person():
    ot_records = OtRecord.query.filter_by(staff=current_user) \
        .filter(OtRecord.canceled_at == None) \
        .filter(OtRecord.round_id != None) \
        .filter(OtRoundRequest.approval_at != None) \
        .filter(OtRoundRequest.verified_at != None).all()
    records = [record.list_records() for record in ot_records]
    return render_template('ot/summary_each_person.html', records=records)


@ot.route('/admin/announcements/<int:announcement_id>/eligible-staff')
@login_required
def view_eligible_staff(announcement_id):
    announcement = OtPaymentAnnounce.query.get(announcement_id)
    return render_template('ot/eligible_staff_list.html', announcement=announcement)


@ot.route('/admin/announcements/<int:announcement_id>/documents')
@login_required
def view_documents(announcement_id):
    announcement = OtPaymentAnnounce.query.get(announcement_id)
    return render_template('ot/documents_list.html', announcement=announcement)


@ot.route('/records/monthly')
@login_required
def view_monthly_records():
    return render_template('ot/staff_calendar.html')


@ot.route('/admin/announcements/<int:announcement_id>/staff/<int:staff_id>/records/monthly')
@login_required
@manager_permission.union(secretary_permission).require()
def view_staff_monthly_records(staff_id, announcement_id):
    staff = StaffAccount.query.get(staff_id)
    return render_template('ot/staff_admin_records.html',
                           staff=staff, announcement_id=announcement_id)


@ot.route('/admin/announcements/<int:announcement_id>/shifts')
@login_required
@manager_permission.union(secretary_permission).require()
def view_shifts(announcement_id):
    announcement = OtPaymentAnnounce.query.get_or_404(announcement_id)
    return render_template('ot/all_staff_calendar.html',
                           announcement_id=announcement_id,
                           announcement=announcement,
                           signatories=announcement.signatories,
                           work_at_org_id=announcement.org_id)


@ot.route('/api/announcements/<int:announcement_id>/ot_shifts')
@login_required
@manager_permission.union(secretary_permission).require()
def get_ot_shifts(announcement_id):
    cal_start = request.args.get('start')
    cal_end = request.args.get('end')
    if cal_start:
        cal_start = parser.isoparse(cal_start)
        cal_start = cal_start.astimezone(localtz)
    if cal_end:
        cal_end = parser.isoparse(cal_end)
        cal_end = cal_end.astimezone(localtz)
    all_shifts = []
    text_color = '#000000'
    for shift in OtShift.query.filter(OtShift.datetime.op('&&')
                                          (DateTimeRange(lower=cal_start,
                                                         upper=cal_end,
                                                         bounds='[]'))) \
            .filter(OtShift.timeslot.has(announcement_id=announcement_id)):
        shift = {
            'title': u'{} คน'.format(len(shift.records)),
            'start': shift.datetime.lower.isoformat(),
            'end': shift.datetime.upper.isoformat(),
            'borderColor': '#000000',
            'backgroundColor': shift.timeslot.color,
            'textColor': text_color,
            'id': shift.id,
        }
        all_shifts.append(shift)
    return jsonify(all_shifts)


@ot.route('/api/ot_records')
@login_required
def get_ot_records():
    cal_start = request.args.get('start')
    cal_end = request.args.get('end')
    if cal_start:
        cal_start = parser.isoparse(cal_start)
    if cal_end:
        cal_end = parser.isoparse(cal_end)
    all_records = []
    text_color = '#000000'
    for shift in OtShift.query.filter(OtShift.datetime.op('&&')
                                          (DateTimeRange(lower=cal_start,
                                                         upper=cal_end,
                                                         bounds='[]'))):
        for record in shift.records:
            if record.staff == current_user:
                start = localtz.localize(record.shift.datetime.lower)
                end = localtz.localize(record.shift.datetime.upper)
                rec = {
                    'title': record.compensation.work_at_org.display_name[:30]
                    if len(record.compensation.work_at_org.display_name) > 30
                    else record.compensation.work_at_org.display_name,
                    'start': start.isoformat(),
                    'end': end.isoformat(),
                    'borderColor': '#000000',
                    'backgroundColor': record.shift.timeslot.color,
                    'textColor': text_color,
                    'id': record.id,
                }
                all_records.append(rec)
    return jsonify(all_records)


# TODO: deprecate this view, use get_all_ot_records_table instead
@ot.route('/api/announcement_id/<int:announcement_id>/ot-records/table')
@login_required
def get_ot_records_table(announcement_id, datetimefmt='%d-%m-%Y %-H:%M'):
    cal_start = request.args.get('start')
    cal_end = request.args.get('end')
    download = request.args.get('download')
    if cal_start:
        cal_start = parser.isoparse(cal_start)
    if cal_end:
        cal_end = parser.isoparse(cal_end)
    all_records = []
    login_pairs = []
    cal_daterange = DateTimeRange(lower=cal_start, upper=cal_end, bounds='[]')
    logins = StaffWorkLogin.query.filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) >= cal_start) \
        .filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) <= cal_end) \
        .filter_by(staff=current_user).order_by(StaffWorkLogin.id).all()

    i = 0
    while i < len(logins):
        if not logins[i].end_datetime:
            _pair = login_tuple(logins[i].staff_id,
                                logins[i].start_datetime.astimezone(localtz),
                                logins[i + 1].start_datetime.astimezone(localtz),
                                logins[i].id,
                                logins[i + 1].id,
                                )
            i += 1
        else:
            _pair = login_tuple(logins[i].staff_id,
                                logins[i].start_datetime.astimezone(localtz),
                                logins[i].end_datetime.astimezone(localtz),
                                logins[i].id,
                                logins[i].id,
                                )
        login_pairs.append(_pair)
        i += 1
    if cal_end and cal_start:
        for shift in OtShift.query.filter(OtShift.datetime.op('&&')(cal_daterange)) \
                .filter(OtShift.timeslot.has(announcement_id=announcement_id)):
            for record in shift.records:
                if record.staff == current_user:
                    shift_start = localtz.localize(record.shift.datetime.lower)
                    shift_end = localtz.localize(record.shift.datetime.upper)
                    overlapped_logins = []
                    overlapped_logouts = []
                    late_mins = []
                    payments = []
                    for _pair in login_pairs:
                        delta_start = _pair.start - shift_start
                        delta_minutes = divmod(delta_start.total_seconds(), 60)
                        if -90 < delta_minutes[0] < 40:
                            overlapped_logins.append(f'{_pair.start.strftime(datetimefmt)}')
                            overlapped_logouts.append(f'{_pair.end.strftime(datetimefmt)}')
                            late_mins.append(str(delta_minutes[0]))
                            if delta_minutes[0] > 0:
                                total_pay = record.calculate_total_pay(record.total_shift_minutes - delta_minutes[0])
                            else:
                                total_pay = record.calculate_total_pay(record.total_shift_minutes)
                            payments.append(total_pay)

                    rec = {
                        'staff': f'{record.staff.fullname}',
                        'title': '{}'.format(record.compensation.ot_job_role),
                        'start': shift_start.isoformat(),
                        'end': shift_end.isoformat(),
                        'id': record.id,
                        'checkins': ','.join(overlapped_logins),
                        'checkouts': ','.join(overlapped_logouts),
                        'late': ','.join([str(m) for m in late_mins]),
                        'payment': ','.join([f'{p:.2f}' for p in payments])
                    }
                    all_records.append(rec)

    if download == 'yes':
        df = pd.DataFrame(all_records)
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name=f'{cal_start.strftime("%Y-%m-%d")}_ot_records.xlsx')

    return jsonify({'data': all_records})


def convert_time_format(time):
    if pd.isna(time):
        return None
    else:
        hours, minutes = divmod(time, 60)
        if hours > 0 or minutes > 0:
            return f'{int(hours)}:{minutes:02.0f}'
        else:
            return None


def write_ot_report_workbook(writer, records_df, format='timesheet'):
    total_work_minutes = records_df.groupby(['fullname', 'sap'])['work_minutes'].sum()
    total_work_minutes.apply(convert_time_format).to_excel(writer, sheet_name='total_minutes')

    total_payment = records_df.groupby(['fullname', 'sap'])['payment'].sum()
    total_payment.to_excel(writer, sheet_name='total_payment')

    renamed_df = records_df.copy()
    if 'staff' in renamed_df.columns:
        del renamed_df['staff']
    renamed_df = renamed_df.rename(columns={
        'sap': 'รหัสบุคคล',
        'fullname': 'ชื่อ',
        'position': 'ตำแหน่งงาน',
        'startDate': 'วันที่',
        'work_minutes': 'เวลาทำงาน',
        'rate': 'อัตรา',
        'timeslot': 'ช่วงเวลา'
    })

    timesheet = renamed_df[['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน', 'อัตรา', 'start', 'end', 'checkins', 'checkouts',
                            'late_checkin_display', 'late_minutes', 'early_checkout_display', 'early_minutes',
                            'เวลาทำงาน', 'payment']]

    if format == 'report':
        summary = renamed_df.pivot_table(['เวลาทำงาน', 'payment'],
                                         ['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน', 'ช่วงเวลา', 'อัตรา'],
                                         'วันที่',
                                         margins=True,
                                         aggfunc='sum')
        summary['ค่าตอบแทน'] = summary[[c for c in summary.columns if c[0] == 'payment' and c[1] != 'All']].sum(axis=1)
        summary = summary[['เวลาทำงาน', 'ค่าตอบแทน']]
        summary['ค่าตอบแทน'] = summary['ค่าตอบแทน'].map(lambda x: round(x, 2))
        summary['เวลาทำงาน'] = summary['เวลาทำงาน'].applymap(convert_time_format)
        summary.to_excel(writer, sheet_name='summary_report')

    timesheet.to_excel(writer, sheet_name='timesheet')


THAI_MONTHS = [
    'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
    'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
]

THAI_WEEKDAYS = ['จ', 'อ', 'พ', 'พฤ', 'ศ', 'ส', 'อา']


def format_buddhist_date_range(start_date, end_date):
    start_text = f'{start_date.day} {THAI_MONTHS[start_date.month - 1]} {start_date.year + 543}'
    end_text = f'{end_date.day} {THAI_MONTHS[end_date.month - 1]} {end_date.year + 543}'
    return f'{start_text} - {end_text}'


def build_custom_ot_report_workbook(records_df, cal_start, cal_end, selected_signatory=None):
    report_df = records_df[records_df['payment'].notna()].copy()
    workbook = Workbook()
    workbook.remove(workbook.active)

    renamed_df = report_df.copy()
    if 'staff' in renamed_df.columns:
        del renamed_df['staff']
    renamed_df = renamed_df.rename(columns={
        'sap': 'รหัสบุคคล',
        'fullname': 'ชื่อ',
        'position': 'ตำแหน่งงาน',
        'startDate': 'วันที่',
        'work_minutes': 'เวลาทำงาน',
        'rate': 'อัตรา',
        'timeslot': 'ช่วงเวลา',
        'start': 'เวลาเริ่มปฏิบัติงาน',
        'end': 'เวลาเลิกปฏิบัติงาน',
        'checkins': 'เวลาเข้างานจริง',
        'checkouts': 'เวลาออกงานจริง',
        'payment': 'จำนวนเงินที่ได้รับ',
    })

    write_total_minutes_custom_sheet(workbook, report_df)
    write_total_payment_custom_sheet(workbook, report_df)
    write_summary_report_custom_sheet(workbook, renamed_df, cal_end)
    write_timesheet_custom_sheet(workbook, renamed_df)
    write_finance_form_custom_sheet(workbook, renamed_df, cal_start, cal_end, selected_signatory=selected_signatory)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def normalize_ot_report_workbook(output):
    return output


def write_total_minutes_custom_sheet(workbook, report_df):
    sheet = workbook.create_sheet('total_minutes  ไม่แก้ไข')
    sheet.append(['fullname', 'sap', 'work_minutes'])
    total_work_minutes = report_df.groupby(['fullname', 'sap'])['work_minutes'].sum().reset_index()
    for row in total_work_minutes.itertuples(index=False):
        sheet.append([row.fullname, row.sap, convert_time_format(row.work_minutes)])


def write_total_payment_custom_sheet(workbook, report_df):
    sheet = workbook.create_sheet('total_payment')
    sheet.append(['ชื่อ - สกุล', 'รหัส sap', 'จำนวนเงินที่ได้รับ'])
    total_payment = report_df.groupby(['fullname', 'sap'])['payment'].sum().reset_index()
    for row in total_payment.itertuples(index=False):
        sheet.append([row.fullname, row.sap, round(row.payment, 2)])
    total_row = sheet.max_row + 1
    sheet.cell(row=total_row, column=3).value = f'=SUM(C2:C{total_row - 1})'


def write_summary_report_custom_sheet(workbook, renamed_df, cal_end):
    sheet = workbook.create_sheet('summary_report ไม่แก้ไข')

    date_values = sorted(renamed_df['วันที่'].dropna().unique())
    start_col = 6
    all_col = start_col + len(date_values)
    payment_col = all_col + 1
    sheet.cell(row=1, column=start_col).value = 'เวลาทำงาน'
    sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=all_col)
    sheet.cell(row=1, column=payment_col).value = 'ค่าตอบแทน'
    sheet.cell(row=2, column=5).value = 'วัน'
    sheet.cell(row=3, column=5).value = 'วันที่'

    for offset, date_value in enumerate(date_values):
        current_date = parser.parse(date_value).date()
        sheet.cell(row=2, column=start_col + offset).value = THAI_WEEKDAYS[current_date.weekday()]
        sheet.cell(row=3, column=start_col + offset).value = current_date.day
    sheet.cell(row=3, column=all_col).value = 'All'

    for col_no, header in enumerate(['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน', 'ช่วงเวลา', 'อัตรา'], start=1):
        sheet.cell(row=4, column=col_no).value = header

    summary = renamed_df.pivot_table(['เวลาทำงาน', 'จำนวนเงินที่ได้รับ'],
                                     ['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน', 'ช่วงเวลา', 'อัตรา'],
                                     'วันที่',
                                     margins=True,
                                     aggfunc='sum')
    payment_cols = [c for c in summary.columns if c[0] == 'จำนวนเงินที่ได้รับ' and c[1] != 'All']
    summary['ค่าตอบแทน'] = summary[payment_cols].sum(axis=1)
    summary = summary[['เวลาทำงาน', 'ค่าตอบแทน']]

    row_no = 5
    current_group_start = None
    current_group_key = None
    for index_values, row_values in summary.iterrows():
        name, sap, position, timeslot, rate = index_values
        is_total_row = name == 'All'
        if is_total_row:
            if current_group_start is not None and row_no - 1 > current_group_start:
                for merge_col in range(1, 4):
                    sheet.merge_cells(start_row=current_group_start, start_column=merge_col,
                                      end_row=row_no - 1, end_column=merge_col)
            current_group_start = None
            current_group_key = None
            sheet.cell(row=row_no, column=1).value = 'All'
        else:
            group_key = (name, sap, position)
            if current_group_key is None:
                current_group_start = row_no
                current_group_key = group_key
            elif group_key != current_group_key:
                if row_no - 1 > current_group_start:
                    for merge_col in range(1, 4):
                        sheet.merge_cells(start_row=current_group_start, start_column=merge_col,
                                          end_row=row_no - 1, end_column=merge_col)
                current_group_start = row_no
                current_group_key = group_key

            sheet.cell(row=row_no, column=1).value = name
            sheet.cell(row=row_no, column=2).value = sap
            sheet.cell(row=row_no, column=3).value = position
            sheet.cell(row=row_no, column=4).value = timeslot
            sheet.cell(row=row_no, column=5).value = rate

        for offset, date_value in enumerate(date_values):
            value = row_values.get(('เวลาทำงาน', date_value))
            sheet.cell(row=row_no, column=start_col + offset).value = convert_time_format(value)
        sheet.cell(row=row_no, column=all_col).value = convert_time_format(row_values.get(('เวลาทำงาน', 'All')))
        payment_value = row_values.get(('ค่าตอบแทน', ''))
        if pd.notna(payment_value):
            sheet.cell(row=row_no, column=payment_col).value = round(payment_value, 2)
        row_no += 1

    if current_group_start is not None and row_no - 2 > current_group_start:
        for merge_col in range(1, 4):
            sheet.merge_cells(start_row=current_group_start, start_column=merge_col,
                              end_row=row_no - 2, end_column=merge_col)


def write_timesheet_custom_sheet(workbook, renamed_df):
    sheet = workbook.create_sheet('timesheet')
    columns = ['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน', 'อัตรา', 'เวลาเริ่มปฏิบัติงาน', 'เวลาเลิกปฏิบัติงาน',
               'เวลาเข้างานจริง', 'เวลาออกงานจริง', 'late_checkin_display', 'late_minutes',
               'early_checkout_display', 'early_minutes', 'จำนวนเวลาปฏิบัติงาน', 'จำนวนเงินที่ได้รับ']
    sheet.append(columns)

    export_df = renamed_df.copy()
    export_df['จำนวนเวลาปฏิบัติงาน'] = export_df['เวลาทำงาน']
    for row in export_df[columns].itertuples(index=False):
        sheet.append(list(row))


def write_finance_form_custom_sheet(workbook, renamed_df, cal_start, cal_end, selected_signatory=None):
    sheet = workbook.create_sheet('ฟอร์มที่ต้องส่งให้การเงิน')
    row_no = 1
    date_range_text = format_buddhist_date_range(cal_start.date(), cal_end.date())
    prepared_by_name = _signatory_display_name(selected_signatory)
    controller_name = _signer_display_name(selected_signatory)
    controller_position = _signer_display_position(selected_signatory)

    grouped = renamed_df.groupby(['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน'], sort=True)
    for group_key in grouped.groups:
        fullname, sap, position = group_key
        staff_df = grouped.get_group(group_key)

        sheet.cell(row=row_no, column=1).value = 'ใบลงเวลา และรายงานผลการปฏิบัติงานนอกเวลาราชการ   ประจำวันที่  '
        sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=4)
        sheet.cell(row=row_no, column=5).value = date_range_text
        sheet.merge_cells(start_row=row_no, start_column=5, end_row=row_no, end_column=7)

        row_no += 1
        sheet.cell(row=row_no, column=1).value = fullname
        sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)
        sheet.cell(row=row_no, column=3).value = 'ตำแหน่งงาน'
        sheet.cell(row=row_no, column=4).value = position
        sheet.cell(row=row_no, column=5).value = 'รหัสบุคคล'
        sheet.cell(row=row_no, column=6).value = sap

        row_no += 1
        headers = ['เวลาเริ่มปฏิบัติงาน', 'เวลาเลิกปฏิบัติงาน', 'อัตราจ่าย', 'เวลาเข้างานจริง',
                   'เวลาออกงานจริง', 'จำนวนเวลาปฏิบัติงาน', 'จำนวนเงินที่ได้รับ']
        for col_no, header in enumerate(headers, start=1):
            sheet.cell(row=row_no, column=col_no).value = header

        for record in staff_df[['เวลาเริ่มปฏิบัติงาน', 'เวลาเลิกปฏิบัติงาน', 'อัตรา', 'เวลาเข้างานจริง',
                                'เวลาออกงานจริง', 'เวลาทำงาน', 'จำนวนเงินที่ได้รับ']].itertuples(index=False):
            row_no += 1
            record_values = list(record)
            record_values[5] = int(record_values[5]) if pd.notna(record_values[5]) else None
            record_values[6] = round(record_values[6], 2) if pd.notna(record_values[6]) else None
            for col_no, value in enumerate(record_values, start=1):
                sheet.cell(row=row_no, column=col_no).value = value

        row_no += 1
        sheet.cell(row=row_no, column=1).value = 'รวมทั้งสิ้น'
        sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=5)
        sheet.cell(row=row_no, column=6).value = convert_time_format(staff_df['เวลาทำงาน'].sum())
        sheet.cell(row=row_no, column=7).value = round(staff_df['จำนวนเงินที่ได้รับ'].sum(), 2)
        row_no += 2

        sheet.cell(row=row_no, column=5).value = 'ข้าพเจ้าขอรับรองว่ามีการปฏิบัติงาน ตามวันเวลาดังกล่าวจริง '
        sheet.merge_cells(start_row=row_no, start_column=5, end_row=row_no, end_column=7)

        row_no += 1
        sheet.cell(row=row_no, column=1).value = f'({fullname})'
        sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)
        sheet.cell(row=row_no, column=3).value = ' ผู้ปฏิบัติงาน'

        row_no += 2
        sheet.cell(row=row_no, column=1).value = f'({prepared_by_name})' if prepared_by_name else ''
        sheet.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=2)
        sheet.cell(row=row_no, column=3).value = 'ผู้จัดทำ'
        sheet.cell(row=row_no, column=5).value = f'({controller_name})' if controller_name else ''
        sheet.merge_cells(start_row=row_no, start_column=5, end_row=row_no, end_column=7)

        row_no += 1
        sheet.cell(row=row_no, column=5).value = controller_position or ''
        sheet.merge_cells(start_row=row_no, start_column=5, end_row=row_no, end_column=7)

        row_no += 1
        sheet.cell(row=row_no, column=5).value = 'ผู้ควบคุมการปฏิบัติงาน'
        sheet.merge_cells(start_row=row_no, start_column=4, end_row=row_no, end_column=7)

        row_no += 3


def _signatory_display_name(signatory):
    if not signatory:
        return ''
    if signatory.report_creator_staff:
        return signatory.report_creator_staff.fullname
    return ''


def _signatory_display_position(signatory):
    if not signatory:
        return ''
    if signatory.report_creator_position:
        return signatory.report_creator_position
    if signatory.report_creator_staff and signatory.report_creator_staff.personal_info:
        return signatory.report_creator_staff.personal_info.position or ''
    return ''


def _signer_display_name(signatory):
    if not signatory:
        return ''
    if signatory.signer_staff:
        return signatory.signer_staff.fullname
    return ''


def _signer_display_position(signatory):
    if not signatory:
        return ''
    if signatory.signer_position:
        return signatory.signer_position
    if signatory.signer_staff and signatory.signer_staff.personal_info:
        return signatory.signer_staff.personal_info.position or ''
    return ''


def build_finance_pdf(records_df, cal_start, cal_end, selected_signatory=None):
    report_df = records_df[records_df['payment'].notna()].copy()
    renamed_df = report_df.copy()
    if 'staff' in renamed_df.columns:
        del renamed_df['staff']
    renamed_df = renamed_df.rename(columns={
        'sap': 'รหัสบุคคล',
        'fullname': 'ชื่อ',
        'position': 'ตำแหน่งงาน',
        'rate': 'อัตรา',
        'start': 'เวลาเริ่มปฏิบัติงาน',
        'end': 'เวลาเลิกปฏิบัติงาน',
        'checkins': 'เวลาเข้างานจริง',
        'checkouts': 'เวลาออกงานจริง',
        'work_minutes': 'เวลาทำงาน',
        'payment': 'จำนวนเงินที่ได้รับ',
    })

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('OtPdfTitle', parent=styles['Normal'], fontName='SarabunBold', fontSize=18,
                                 leading=20, alignment=TA_CENTER)
    normal_style = ParagraphStyle('OtPdfNormal', parent=styles['Normal'], fontName='Sarabun', fontSize=13,
                                  leading=15, alignment=TA_LEFT)
    center_style = ParagraphStyle('OtPdfCenter', parent=normal_style, alignment=TA_CENTER)
    right_style = ParagraphStyle('OtPdfRight', parent=normal_style, alignment=TA_RIGHT)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    story = []
    date_range_text = format_buddhist_date_range(cal_start.date(), cal_end.date())
    grouped = renamed_df.groupby(['ชื่อ', 'รหัสบุคคล', 'ตำแหน่งงาน'], sort=True)

    for section_index, group_key in enumerate(grouped.groups):
        fullname, sap, position = group_key
        staff_df = grouped.get_group(group_key)

        if section_index > 0:
            story.append(PageBreak())

        story.append(Paragraph('ใบลงเวลา และรายงานผลการปฏิบัติงานนอกเวลาราชการ', title_style))
        story.append(Paragraph(f'ประจำวันที่ {date_range_text}', center_style))
        story.append(Spacer(1, 4 * mm))

        info_table = Table([
            [Paragraph(fullname, normal_style), Paragraph('ตำแหน่งงาน', center_style),
             Paragraph(position, normal_style), Paragraph('รหัสบุคคล', center_style), Paragraph(str(sap), normal_style)]
        ], colWidths=[55 * mm, 22 * mm, 45 * mm, 22 * mm, 28 * mm])
        info_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 3 * mm))

        header_row = [
            Paragraph('เวลาเริ่มปฏิบัติงาน', center_style),
            Paragraph('เวลาเลิกปฏิบัติงาน', center_style),
            Paragraph('อัตราจ่าย', center_style),
            Paragraph('เวลาเข้างานจริง', center_style),
            Paragraph('เวลาออกงานจริง', center_style),
            Paragraph('จำนวนเวลาปฏิบัติงาน', center_style),
            Paragraph('จำนวนเงินที่ได้รับ', center_style),
        ]
        rows = [header_row]
        for record in staff_df[['เวลาเริ่มปฏิบัติงาน', 'เวลาเลิกปฏิบัติงาน', 'อัตรา', 'เวลาเข้างานจริง',
                                'เวลาออกงานจริง', 'เวลาทำงาน', 'จำนวนเงินที่ได้รับ']].itertuples(index=False):
            rows.append([
                Paragraph(str(record[0] or ''), normal_style),
                Paragraph(str(record[1] or ''), normal_style),
                Paragraph(str(record[2] or ''), center_style),
                Paragraph(str(record[3] or ''), normal_style),
                Paragraph(str(record[4] or ''), normal_style),
                Paragraph(str(convert_time_format(record[5]) or ''), center_style),
                Paragraph(f'{record[6]:,.2f}' if pd.notna(record[6]) else '', right_style),
            ])

        total_minutes = convert_time_format(staff_df['เวลาทำงาน'].sum()) or ''
        total_payment = staff_df['จำนวนเงินที่ได้รับ'].sum()
        rows.append([
            Paragraph('รวมทั้งสิ้น', normal_style), '', '', '', '',
            Paragraph(total_minutes, center_style),
            Paragraph(f'{total_payment:,.2f}', right_style),
        ])

        detail_table = Table(rows, repeatRows=1,
                             colWidths=[28 * mm, 28 * mm, 18 * mm, 28 * mm, 28 * mm, 22 * mm, 24 * mm])
        detail_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('SPAN', (0, -1), (4, -1)),
        ]))
        story.append(detail_table)
        story.append(Spacer(1, 5 * mm))

        prepared_by_name = _signatory_display_name(selected_signatory)
        controller_name = _signer_display_name(selected_signatory)
        controller_position = _signer_display_position(selected_signatory)

        footer_rows = [
            [
                Paragraph(f'({fullname})<br/>ผู้ปฏิบัติงาน', center_style),
                Paragraph('ข้าพเจ้าขอรับรองว่ามีการปฏิบัติงาน ตามวันเวลาดังกล่าวจริง', center_style),
            ],
            [
                Paragraph(f'({prepared_by_name})<br/>ผู้จัดทำ' if prepared_by_name else '<br/>ผู้จัดทำ', center_style),
                Paragraph(
                    f'({controller_name})<br/>{controller_position} ผู้ควบคุมการปฏิบัติงาน'
                    if controller_name or controller_position
                    else '<br/><br/>ผู้ควบคุมการปฏิบัติงาน',
                    center_style,
                ),
            ],
        ]
        footer_table = Table(footer_rows, colWidths=[90 * mm, 90 * mm], rowHeights=[18 * mm, 28 * mm])
        footer_table.setStyle(TableStyle([
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
            ('VALIGN', (0, 0), (-1, 0), 'TOP'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LINEBELOW', (0, 0), (-1, 0), 0, colors.white),
            ('LINEBELOW', (0, 1), (-1, 1), 0, colors.white),
        ]))
        story.append(footer_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def humanized_work_time(work_time_minutes):
    hours, minutes = divmod(work_time_minutes, 60)
    h = f'{hours:.0f}h'
    m = f'{minutes:.0f}m'
    if hours and minutes:
        return f'{h}:{m}'
    elif hours:
        return h
    else:
        return m


def _build_checkin_pairs(checkin_query):
    """Normalize raw login rows into reusable check-in/check-out pairs."""
    logins = defaultdict(list)
    for checkin in checkin_query.order_by(StaffWorkLogin.start_datetime):
        logins[checkin.staff_id].append(checkin)

    checkin_pairs = defaultdict(list)
    for checkin_staff_id, checkins in logins.items():
        i = 0
        while i < len(checkins):
            curr_start = checkins[i].start_datetime.astimezone(localtz).replace(second=0, microsecond=0)
            if checkins[i].end_datetime:
                curr_end = checkins[i].end_datetime.astimezone(localtz).replace(second=0, microsecond=0)
                pair = login_tuple(checkin_staff_id, curr_start, curr_end, checkins[i].id, checkins[i].id)
                checkin_pairs[checkin_staff_id].append(pair)
            else:
                try:
                    next_start = checkins[i + 1].start_datetime.astimezone(localtz).replace(second=0, microsecond=0)
                except IndexError:
                    pair = login_tuple(checkin_staff_id, curr_start, None, checkins[i].id, None)
                    checkin_pairs[checkin_staff_id].append(pair)
                else:
                    _delta_days = (next_start.date() - curr_start.date()).days
                    if _delta_days == 1 and (
                        curr_start.time() >= time(18, 0) or next_start.time() <= time(4, 0)
                    ):
                        # Split a cross-midnight open sequence at midnight so the overnight shift
                        # can use the midnight scan and the next day keeps its own check-in.
                        _d = curr_start + timedelta(days=1)
                        midnight1 = _d.replace(hour=0, minute=0, second=0, microsecond=0)
                        midnight2 = next_start.replace(hour=0, minute=0, second=0, microsecond=0)
                        pair = login_tuple(checkin_staff_id, curr_start, midnight1, checkins[i].id, None)
                        pair2 = login_tuple(checkin_staff_id, midnight2, next_start, None, checkins[i + 1].id)
                        checkin_pairs[checkin_staff_id].append(pair)
                        checkin_pairs[checkin_staff_id].append(pair2)
                    elif _delta_days == 0 and checkins[i + 1].end_datetime is None:
                        pair = login_tuple(checkin_staff_id, curr_start, next_start, checkins[i].id, checkins[i + 1].id)
                        checkin_pairs[checkin_staff_id].append(pair)
                    else:
                        # Same-day open rows are left open so they do not consume the next scan.
                        pair = login_tuple(checkin_staff_id, curr_start, None, checkins[i].id, None)
                        checkin_pairs[checkin_staff_id].append(pair)
            i += 1

    return checkin_pairs


def _compute_work_minutes(record, shift_start, shift_end, pair):
    """Compute attendance timing and pay for a single shift pairing."""
    checkin = pair.start.isoformat() if not request.args.get('download') else pair.start.strftime('%Y-%m-%d %H:%M:%S')
    start_delta_minutes = divmod((pair.start - shift_start).total_seconds(), 60)

    if pair.end:
        checkout = pair.end.isoformat() if not request.args.get('download') else pair.end.strftime('%Y-%m-%d %H:%M:%S')
        if pair.end < shift_start:
            return None
    else:
        checkout = None
    if record.compensation.per_period:
        if pair.end is None:
            return {
                'checkin': checkin,
                'checkout': checkout,
                'checkin_late_minutes': 0,
                'checkout_early_minutes': 0,
                'total_work_minutes': None,
                'total_pay': None,
                'missing_checkout': True,
            }
        checkin_late_minutes = 0
        checkout_early_minutes = 0
        total_work_minutes = record.total_shift_minutes
        total_pay = round(record.calculate_total_pay(total_work_minutes), 2)
        return {
            'checkin': checkin,
            'checkout': checkout,
            'checkin_late_minutes': checkin_late_minutes,
            'checkout_early_minutes': checkout_early_minutes,
            'total_work_minutes': total_work_minutes,
            'total_pay': total_pay,
            'missing_checkout': False,
        }

    if pair.end:
        if pair.end < shift_end:
            delta_end = shift_end - pair.end
            end_delta_minutes = divmod(delta_end.total_seconds(), 60)
            print('end_delta_minutes:', end_delta_minutes, delta_end)
        else:
            end_delta_minutes = (0, 0)
    else:
        end_delta_minutes = (0, 0)

    checkin_late_minutes = 0 if start_delta_minutes[0] < 0 else start_delta_minutes[0]
    checkout_early_minutes = 0 if end_delta_minutes[0] < 0 else end_delta_minutes[0]
    if checkin_late_minutes > 0 or checkout_early_minutes > 0:
        total_work_minutes = record.total_shift_minutes - checkin_late_minutes - checkout_early_minutes
        total_pay = round(record.calculate_total_pay(total_work_minutes), 2)
    else:
        total_pay = round(record.calculate_total_pay(record.total_shift_minutes), 2)
        total_work_minutes = record.total_shift_minutes

    if pair.end is None:
        return {
            'checkin': checkin,
            'checkout': checkout,
            'checkin_late_minutes': 0 if start_delta_minutes[0] < 0 else start_delta_minutes[0],
            'checkout_early_minutes': 0,
            'total_work_minutes': None,
            'total_pay': None,
            'missing_checkout': True,
        }

    return {
        'checkin': checkin,
        'checkout': checkout,
        'checkin_late_minutes': checkin_late_minutes,
        'checkout_early_minutes': checkout_early_minutes,
        'total_work_minutes': total_work_minutes,
        'total_pay': total_pay,
        'missing_checkout': False,
    }


def _build_ot_record_row(record, shift_start, shift_end, announcement_id, staff_id, download):
    """Build the base JSON row returned for one OT record."""
    return {
        'fullname': f'{record.staff.fullname}',
        'sap': f'{record.staff.personal_info.sap_id}',
        'timeslot': f'{record.compensation.time_slot}' if record.compensation else '-',
        'staff': f'{record.staff.fullname}' if staff_id else f'''<a href="{url_for('ot.view_staff_monthly_records', staff_id=record.staff_account_id, announcement_id=announcement_id)}">{record.staff.fullname}</a>''',
        'start': shift_start.isoformat() if not download else shift_start.strftime('%Y-%m-%d %H:%M:%S'),
        'end': shift_end.isoformat() if not download else shift_end.strftime('%Y-%m-%d %H:%M:%S'),
        'id': record.id,
        'checkin_staff_id': record.staff_account_id,
        'checkin_id': None,
        'checkout_id': None,
        'checkins': None,
        'checkouts': None,
        'late_checkin_display': None,
        'late_minutes': None,
        'early_minutes': None,
        'early_checkout_display': None,
        'payment': None,
        'work_minutes': None,
        'work_minutes_display': None,
        'missing_checkout': False,
        'position': record.compensation.ot_job_role.role if record.compensation else '-',
        'rate': record.compensation.rate if record.compensation else '-',
        'startDate': shift_start.strftime('%Y/%m/%d'),
        'endDate': shift_end.strftime('%Y/%m/%d'),
        'workAt': record.compensation.work_at_org.display_name,
    }

@ot.route('/api/announcement_id/<int:announcement_id>/staff/<int:staff_id>/ot-schedule')
@ot.route('/api/announcement_id/<int:announcement_id>/staff/ot-schedule')
@login_required
def get_all_ot_schedule(announcement_id=None, staff_id=None):
    cal_start = request.args.get('start')
    cal_end = request.args.get('end')
    if cal_start:
        cal_start = parser.isoparse(cal_start)
        cal_start = cal_start.astimezone(localtz)
    if cal_end:
        cal_end = parser.isoparse(cal_end)
        cal_end = cal_end.astimezone(localtz)
    cal_daterange = DateTimeRange(lower=cal_start, upper=cal_end, bounds='[]')
    shift_query = OtShift.query.filter(OtShift.datetime.op('&&')(cal_daterange))
    all_records = []

    for shift in shift_query.order_by(OtShift.datetime):
        for record in shift.records:
            if staff_id and record.staff_account_id != staff_id:
                continue
            shift_start = localtz.localize(record.shift.datetime.lower)
            shift_end = localtz.localize(record.shift.datetime.upper)

            rec = {
                'fullname': f'{record.staff.fullname}',
                'sap': f'{record.staff.personal_info.sap_id}',
                'timeslot': f'{record.compensation.time_slot}' if record.compensation else '-',
                'staff': f'{record.staff.fullname}' if staff_id else f'''<a href="{url_for('ot.view_staff_monthly_records', staff_id=record.staff_account_id, announcement_id=announcement_id)}">{record.staff.fullname}</a>''',
                'start': shift_start.strftime('%Y-%m-%d %H:%M:%S'),
                'end': shift_end.strftime('%Y-%m-%d %H:%M:%S'),
                'id': record.id,
                'position': record.compensation.ot_job_role.role if record.compensation else '-',
                'rate': record.compensation.rate if record.compensation else '-',
                'startDate': shift_start.strftime('%Y/%m/%d'),
                'endDate': shift_end.strftime('%Y/%m/%d'),
                'workAt': record.compensation.work_at_org.display_name,
            }
            all_records.append(rec)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df = pd.DataFrame(all_records)
                schedule = (df.groupby(['fullname', 'sap', 'position', 'timeslot'])['startDate'].count().to_excel(writer, sheet_name='schedule'))
                del df['staff']
                df = df.rename(columns={
                    'sap': 'รหัสบุคคล',
                    'fullname': 'ชื่อ',
                    'position': 'ตำแหน่งงาน',
                    'startDate': 'วันที่',
                    'timeslot': 'ช่วงเวลา'
                })
        output.seek(0)
        if staff_id:
            staff = StaffAccount.query.get(staff_id)
            download_name = f'{staff.email}_{cal_start.strftime("%m-%Y")}_ot_{format}.xlsx'
        else:
            download_name = f'{cal_start.strftime("%m-%Y")}_ot_{format}_all.xlsx'
        return send_file(output, download_name=download_name)


@ot.route('/api/announcement_id/<int:announcement_id>/staff/<int:staff_id>/ot-records/table')
@ot.route('/api/announcement_id/<int:announcement_id>/staff/ot-records/table')
@ot.route('/api/staff/<int:staff_id>/ot-records/table')
@login_required
def get_all_ot_records_table(announcement_id=None, staff_id=None):
    cal_start = request.args.get('start')
    cal_end = request.args.get('end')
    download = request.args.get('download')
    format = request.args.get('format', 'timesheet')
    selected_signatory_id = request.args.get('signatory_id', type=int)
    if _is_external_account():
        if staff_id is not None and staff_id != current_user.id:
            abort(403)
        staff_id = current_user.id
    if cal_start:
        cal_start = parser.isoparse(cal_start)
        cal_start = cal_start.astimezone(localtz)
    if cal_end:
        cal_end = parser.isoparse(cal_end)
        cal_end = cal_end.astimezone(localtz)

    cal_daterange = DateTimeRange(lower=cal_start, upper=cal_end, bounds='[]')
    checkin_query = StaffWorkLogin.query\
        .filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) >= cal_start) \
        .filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) <= cal_end) \

    if staff_id:
        checkin_query = checkin_query.filter_by(staff_id=staff_id)
    checkin_pairs = _build_checkin_pairs(checkin_query)

    shift_query = OtShift.query.filter(OtShift.datetime.op('&&')(cal_daterange))
    if announcement_id:
        shift_query = shift_query.filter(OtShift.timeslot.has(announcement_id=announcement_id))

    all_records = []
    used_checkouts = defaultdict(set)
    ot_record_checkins = {}
    for shift in shift_query.order_by(OtShift.datetime):
        for record in shift.records:
            if staff_id and record.staff_account_id != staff_id:
                continue
            shift_start = localtz.localize(record.shift.datetime.lower)
            shift_end = localtz.localize(record.shift.datetime.upper)
            rec = _build_ot_record_row(record, shift_start, shift_end, announcement_id, staff_id, download)
            ot_record_checkins[record] = 0

            if checkin_pairs[record.staff_account_id]:
                for _pair in checkin_pairs[record.staff_account_id]:
                    # Ignore scan pairs that do not fall inside the shift date window.
                    if _pair.start and _pair.end:
                        if _pair.start.date() != shift_start.date() and _pair.end.date() != shift_end.date():
                            continue

                    # Treat midnight as a real check-in only for midnight-starting shifts.
                    if _pair.start.time() == time(0, 0) and shift_start.time() != _pair.start.time():
                        if _pair.end and _pair.start_id is None:
                            used_checkouts[record.staff_account_id].add(_pair.end.strftime('%Y-%m-%d %H:%M:%S'))
                        continue
                    # Do not reuse a checkout after midnight as a later check-in.
                    if _pair.start.strftime('%Y-%m-%d %H:%M:%S') in used_checkouts[record.staff_account_id]:
                        continue

                    attendance = _compute_work_minutes(record, shift_start, shift_end, _pair)
                    if not attendance:
                        continue

                    checkin = attendance['checkin']
                    checkout = attendance['checkout']
                    checkin_late_minutes = attendance['checkin_late_minutes']
                    checkout_early_minutes = attendance['checkout_early_minutes']
                    total_work_minutes = attendance['total_work_minutes']
                    total_pay = attendance['total_pay']

                    if total_work_minutes is None:
                        rec.update({
                            'checkins': checkin,
                            'checkouts': checkout,
                            'late_checkin_display': f'{humanized_work_time(checkin_late_minutes)}' if checkin_late_minutes else None,
                            'late_minutes': checkin_late_minutes,
                            'early_minutes': checkout_early_minutes,
                            'early_checkout_display': f'{humanized_work_time(checkout_early_minutes)}' if checkout_early_minutes else None,
                            'missing_checkout': True,
                        })
                        if _pair.end and _pair.start_id is None:
                            used_checkouts[record.staff_account_id].add(_pair.end.strftime('%Y-%m-%d %H:%M:%S'))
                        continue

                    if total_work_minutes > 0 and checkin_late_minutes <= MAX_LATE_MINUTES:
                        rec.update({
                            'checkin_staff_id': _pair.staff_id,
                            'checkin_id': _pair.start_id,
                            'checkout_id': _pair.end_id,
                            'checkins': checkin,
                            'checkouts': checkout,
                            'late_checkin_display': f'{humanized_work_time(checkin_late_minutes)}' if checkin_late_minutes else None,
                            'late_minutes': checkin_late_minutes,
                            'early_minutes': checkout_early_minutes,
                            'early_checkout_display': f'{humanized_work_time(checkout_early_minutes)}' if checkout_early_minutes else None,
                            'payment': total_pay,
                            'work_minutes': total_work_minutes,
                            'work_minutes_display': f'{humanized_work_time(total_work_minutes)}' if total_work_minutes else None,
                        })
                        ot_record_checkins[record] += 1
                        if _pair.end and _pair.start_id is None:
                            used_checkouts[record.staff_account_id].add(_pair.end.strftime('%Y-%m-%d %H:%M:%S'))
                        break
            all_records.append(rec)

    if download == 'yes':
        if request.args.get('download_data') == 'counts':
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                missing_checkins = []
                for r, c in ot_record_checkins.items():
                    missing_checkins.append({
                        'record_id': r.id,
                        'staff': r.staff.fullname,
                        'position': r.compensation.ot_job_role.role if r.compensation else '-',
                        'rate': r.compensation.rate if r.compensation else '-',
                        'start': r.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                        'end': r.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                        'count': c
                    })
                df = pd.DataFrame(missing_checkins)
                df.to_excel(writer, sheet_name='counts')
        else:
            df = pd.DataFrame(all_records)
            selected_signatory = None
            if announcement_id and format in ('finance-pdf', 'report'):
                announcement = OtPaymentAnnounce.query.get(announcement_id)
                if announcement:
                    signatories = list(announcement.signatories)
                    if selected_signatory_id:
                        selected_signatory = next((signatory for signatory in signatories
                                                   if signatory.id == selected_signatory_id), None)
                    if not selected_signatory:
                        selected_signatory = next((signatory for signatory in signatories
                                                   if signatory.report_creator_staff), None)
            if format == 'report':
                output = build_custom_ot_report_workbook(df, cal_start, cal_end,
                                                         selected_signatory=selected_signatory)
            elif format == 'finance-pdf':
                output = build_finance_pdf(df, cal_start, cal_end, selected_signatory=selected_signatory)
            else:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    write_ot_report_workbook(writer, df, format=format)
        output.seek(0)
        if staff_id:
            staff = StaffAccount.query.get(staff_id)
            ext = 'pdf' if format == 'finance-pdf' else 'xlsx'
            download_name = f'{staff.email}_{cal_start.strftime("%m-%Y")}_ot_{format}.{ext}'
        else:
            ext = 'pdf' if format == 'finance-pdf' else 'xlsx'
            download_name = f'{cal_start.strftime("%m-%Y")}_ot_{format}_all.{ext}'
        return send_file(output, download_name=download_name)
    return jsonify({'data': all_records})


@ot.route('/api/staff/<int:staff_id>/checkin-records', methods=['GET', 'POST'])
@ot.route('/api/checkin-records/<int:checkin_id>', methods=['DELETE'])
@login_required
def add_checkin_record(staff_id=None, checkin_id=None):
    # TODO: Restrict this endpoint to manager/secretary roles like the admin pages.
    if request.method == 'GET':
        download = request.args.get('download', 'no')
        cal_start = request.args.get('start')
        cal_end = request.args.get('end')
        if cal_start:
            cal_start = parser.isoparse(cal_start)
            cal_start = cal_start.astimezone(localtz)
        if cal_end:
            cal_end = parser.isoparse(cal_end)
            cal_end = cal_end.astimezone(localtz)

        staff = StaffAccount.query.get(staff_id)

        query = StaffWorkLogin.query.filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) >= cal_start) \
            .filter(func.timezone('Asia/Bangkok', StaffWorkLogin.start_datetime) <= cal_end) \
            .filter_by(staff=staff) \
            .order_by(StaffWorkLogin.start_datetime)

        if download == 'yes':
            logins = query.all()
            login_pairs = []
            i = 0
            while i < len(logins):
                if not logins[i].end_datetime:
                    _start = logins[i].start_datetime.astimezone(localtz).strftime('%Y-%m-%d %H:%M:%S')
                    try:
                        _end = logins[i + 1].start_datetime.astimezone(localtz).strftime('%Y-%m-%d %H:%M:%S')
                    except IndexError:
                        _pair = {'checkin': _start, 'checkout': None, 'staff': logins[i].staff.fullname}
                    else:
                        _pair = {'checkin': _start, 'checkout': _end, 'staff': logins[i].staff.fullname}
                        i += 1
                else:
                    _pair = {
                        'staff': logins[i].staff.fullname,
                        'checkin': logins[i].start_datetime.astimezone(localtz).strftime('%Y-%m-%d %H:%M:%S'),
                        'checkout': logins[i].end_datetime.astimezone(localtz).strftime('%Y-%m-%d %H:%M:%S'),
                    }
                login_pairs.append(_pair)
                i += 1
            df = pd.DataFrame(login_pairs)
            output = io.BytesIO()
            df[['staff', 'checkin', 'checkout']].to_excel(output)
            output.seek(0)
            return send_file(output, download_name=f'{cal_start.strftime("%Y-%m-%d")}_ot_checkins.xlsx')
        else:
            all_records = []
            for checkin in query:
                rec = {
                    'staff': staff.fullname,
                    'note': checkin.note,
                    'checkin': checkin.start_datetime.isoformat() if download == 'no' else checkin.start_datetime.strftime(
                        '%Y-%m-%d %H:%M:%S'),
                    'action': f'<a onclick="deleteCheckin({checkin.id})">delete</a>'
                }
                all_records.append(rec)
            return jsonify({'data': all_records})
    elif request.method == 'DELETE':
        checkin = StaffWorkLogin.query.get(checkin_id)
        db.session.delete(checkin)
        db.session.commit()
        return jsonify({'message': 'success'})
    elif request.method == 'POST':
        form = request.form
        checkin_datetime = form.get('checkin-datetime')
        checkin_datetime = arrow.get(datetime.strptime(checkin_datetime, '%d/%m/%Y %H:%M:%S'), 'Asia/Bangkok').datetime
        new_checkin_record = StaffWorkLogin()
        new_checkin_record.staff_id = staff_id
        new_checkin_record.start_datetime = checkin_datetime
        note = form.get('note')
        new_checkin_record.note = note or 'แก้ไข/เพิ่มเติมโดย admin'
        db.session.add(new_checkin_record)
        db.session.commit()
        resp = make_response()
        resp.headers['HX-Trigger'] = 'reload.data'
        return resp
